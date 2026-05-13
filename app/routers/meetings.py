"""Router Tenues & Agapes — CRUD + inscription publique + présences"""
from datetime import date, datetime, timedelta
from typing import Annotated, Optional
import secrets

import csv
import io

from fastapi import APIRouter, Depends, Form, HTTPException, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy import select, func as sql_func
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import (
    require_auth, can_manage_meeting, can_lock_meeting,
)
from app.models.reports import MeetingReport
from app.models.meetings import (
    Meeting, Attendance, AttendanceStatus, DegreeAttended,
    MeetingGrade, MeetingType, MeetingDegree,
    Visitor, MeetingVisitor, MeetingGuest, MeetingWaitlist,
    DietaryRestriction, GuestStatus,
)
from app.models.identity import Member, LodgeFunction
from app.models.lodge import MasonicYear, LodgeSettings, LodgeOffice, MeetingOffice

router = APIRouter(prefix="/meetings", tags=["meetings"])
templates = Jinja2Templates(directory="app/templates")


async def _count_agapes(db: AsyncSession, meeting_id: int) -> int:
    """Compte le total de couverts agapes pour une tenue (membres + visiteurs confirmés)."""
    # Membres
    r1 = await db.execute(
        select(sql_func.sum(Attendance.agape_guests + 1))
        .where(Attendance.meeting_id == meeting_id, Attendance.agape == True)
    )
    m = r1.scalar() or 0
    # Visiteurs confirmés
    r2 = await db.execute(
        select(sql_func.sum(MeetingVisitor.agape_guests + 1))
        .where(MeetingVisitor.meeting_id == meeting_id,
               MeetingVisitor.agape == True,
               MeetingVisitor.status == "CONFIRMED")
    )
    v = r2.scalar() or 0
    return int(m + v)


def _type_label(t) -> str:
    labels = {
        "BLANCHE":      "Tenue blanche",
        "SOLENNELLE":   "Tenue solennelle",
        "INSTRUCTION":  "Tenue d'instruction",
        "INITIATION":   "Initiation",
        "INSTALLATION": "Installation des officiers",
        "ELECTION":     "Élection du Vénérable",
        "PASSAGE":      "Passage au 2e degré",
        "ELEVATION":    "Élévation au 3e degré",
        "FETE":         "Fête maçonnique",
        "EXTRA":        "Tenue extraordinaire",
    }
    v = t.value if hasattr(t, "value") else str(t)
    return labels.get(v, v)


def _grade_label(g: MeetingGrade) -> str:
    return {
        "APPRENTI": "Apprentis",
        "COMPAGNON": "Compagnons",
        "MAITRE": "Maîtres",
        "ALL": "Toutes loges réunies",
    }.get(g, g)


# ── Liste des tenues ──────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def meetings_list(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    year_id: int = 0,
    upcoming_only: str = "0",   # "0" = absent du form (checkbox décochée)
    submitted: str = "0",
):
    user, member = ctx

    # Année maçonnique courante
    year_q = select(MasonicYear).order_by(MasonicYear.is_current.desc(), MasonicYear.start_date.desc())
    years_result = await db.execute(year_q)
    years = years_result.scalars().all()

    current_year = next((y for y in years if y.is_current), years[0] if years else None)
    selected_year = current_year

    if year_id:
        found = next((y for y in years if y.id == year_id), None)
        if found:
            selected_year = found

    # Requête tenues
    q = select(Meeting).order_by(Meeting.meeting_date.desc())
    if selected_year:
        q = q.where(Meeting.masonic_year_id == selected_year.id)
    # Si le form n'a pas encore été soumis → comportement par défaut = à venir
    # Si soumis sans la checkbox → upcoming_only sera absent → "0"
    effective_upcoming = upcoming_only if submitted == "1" else "1"
    if effective_upcoming == "1":
        q = q.where(Meeting.meeting_date >= date.today())
        q = q.order_by(Meeting.meeting_date.asc())

    result = await db.execute(q)
    meetings = result.scalars().all()

    # Nombre de tenues passées (pour afficher le hint quand filtre "à venir" actif)
    past_count = 0
    if effective_upcoming == "1" and selected_year:
        pc_r = await db.execute(
            select(sql_func.count()).where(
                Meeting.masonic_year_id == selected_year.id,
                Meeting.meeting_date < date.today(),
            )
        )
        past_count = pc_r.scalar() or 0

    # Compter les présences pour chaque tenue
    attendance_counts = {}
    if meetings:
        ids = [m.id for m in meetings]
        count_q = select(
            Attendance.meeting_id,
            sql_func.count(Attendance.id)
        ).where(
            Attendance.meeting_id.in_(ids),
            Attendance.status == AttendanceStatus.PRESENT
        ).group_by(Attendance.meeting_id)
        count_result = await db.execute(count_q)
        attendance_counts = dict(count_result.all())

    # Inscription du membre courant
    member_attendances = {}
    if meetings:
        att_q = select(Attendance).where(
            Attendance.meeting_id.in_([m.id for m in meetings]),
            Attendance.member_id == member.id,
        )
        att_result = await db.execute(att_q)
        for att in att_result.scalars().all():
            member_attendances[att.meeting_id] = att

    return templates.TemplateResponse(request, "pages/meetings/list.html", {
        "current_member": member,
        "current_user": user,
        "meetings": meetings,
        "years": years,
        "selected_year": selected_year,
        "upcoming_only": effective_upcoming,
        "attendance_counts": attendance_counts,
        "member_attendances": member_attendances,
        "type_label": _type_label,
        "grade_label": _grade_label,
        "can_manage": can_manage_meeting(member) or user.is_admin,
        "today": date.today(),
        "past_count": past_count,
    })


# ── Détail d'une tenue ────────────────────────────────────────────────────────

@router.get("/{meeting_id}", response_class=HTMLResponse)
async def meeting_detail(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx

    # Charger la tenue avec ses présences
    result = await db.execute(
        select(Meeting)
        .options(
            selectinload(Meeting.attendances).selectinload(Attendance.member),
            selectinload(Meeting.meeting_visitors).selectinload(MeetingVisitor.visitor),
            selectinload(Meeting.meeting_guests),
            selectinload(Meeting.waitlist),
            selectinload(Meeting.degrees),
        )
        .where(Meeting.id == meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404, detail="Tenue introuvable")

    # Présence du membre courant
    my_attendance = next(
        (a for a in meeting.attendances if a.member_id == member.id), None
    )

    # Compter agapes
    agape_count = sum(
        (1 if a.agape else 0) + a.agape_guests
        for a in meeting.attendances
    ) + sum(
        (1 if mv.agape else 0) + mv.agape_guests
        for mv in meeting.meeting_visitors
        if mv.status.value == "CONFIRMED"
    ) + sum(
        1 for g in meeting.meeting_guests
        if g.status.value == "CONFIRMED" and g.agape
    )

    present_count = sum(1 for a in meeting.attendances if a.status.value == "PRESENT")
    excused_count = sum(1 for a in meeting.attendances if a.status.value == "EXCUSED")

    # PV de la tenue
    report_r = await db.execute(
        select(MeetingReport).where(MeetingReport.meeting_id == meeting_id)
    )
    report = report_r.scalar_one_or_none()
    from app.routers.reports import _can_write as _report_can_write, _can_approve as _report_can_approve
    can_write_report = _report_can_write(user, member)
    can_approve_report = _report_can_approve(user, member)

    return templates.TemplateResponse(request, "pages/meetings/detail.html", {
        "current_member": member,
        "current_user": user,
        "meeting": meeting,
        "my_attendance": my_attendance,
        "present_count": present_count,
        "excused_count": excused_count,
        "agape_count": agape_count,
        "type_label": _type_label,
        "grade_label": _grade_label,
        "AttendanceStatus": AttendanceStatus,
        "can_manage": can_manage_meeting(member) or user.is_admin,
        "can_lock": can_lock_meeting(member),
        "registration_url": f"{request.base_url}inscription/{meeting.token}",
        "report": report,
        "can_write_report": can_write_report,
        "can_approve_report": can_approve_report,
    })


# ── Export Excel agapes ───────────────────────────────────────────────────────

@router.get("/{meeting_id}/agapes/export")
async def agapes_export_excel(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    if not (can_manage_meeting(member) or user.is_admin
            or member.lodge_function == LodgeFunction.MAITRE_BANQUETS):
        raise HTTPException(status_code=403, detail="Accès refusé")

    result = await db.execute(
        select(Meeting)
        .options(
            selectinload(Meeting.attendances).selectinload(Attendance.member),
            selectinload(Meeting.meeting_visitors).selectinload(MeetingVisitor.visitor),
            selectinload(Meeting.meeting_guests),
        )
        .where(Meeting.id == meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    diet_labels = {
        "NONE": "—",
        "VEGETARIAN": "Végétarien 🥦",
        "VEGAN": "Vegan 🌱",
        "NO_PORK": "Sans porc 🚫",
        "OTHER": "Régime spécial ⚠",
    }

    # Couleurs
    C_HEADER   = "4A3728"   # brun foncé
    C_FRERE    = "E8F0FE"   # bleu pâle
    C_VISITEUR = "EFF6FF"   # bleu plus pâle
    C_INVITE   = "FFFBEB"   # ambre pâle
    C_SECTION  = "F8FAFC"   # gris très clair
    C_TOTAL    = "FEF3C7"   # ambre léger
    C_DIET     = "FEF9C3"   # jaune pâle

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ── Feuille 1 : Liste nominative ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Liste banquet"
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"] = f"🍽  Agapes du {meeting.meeting_date.strftime('%d/%m/%Y')} — {_type_label(meeting.type)}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sous-titre
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="9CA3AF")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # En-têtes colonnes
    headers = ["Nom", "Prénom", "Catégorie", "Loge / Origine", "Invités (+)", "Régime alimentaire"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[4].height = 22

    row = 5

    def _section_row(ws, label, color, row):
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = label
        cell.font = Font(bold=True, size=9, color="374151")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(indent=1, vertical="center")
        ws.row_dimensions[row].height = 18
        return row + 1

    def _data_row(ws, values, fill_color, row, bold_col=None):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            cell.alignment = Alignment(vertical="center", indent=1)
            if bold_col and col == bold_col:
                cell.font = Font(bold=True)
        ws.row_dimensions[row].height = 18
        return row + 1

    # Membres
    agape_members = sorted([a for a in meeting.attendances if a.agape],
                           key=lambda a: a.member.last_name)
    if agape_members:
        row = _section_row(ws, f"  Membres de la loge ({len(agape_members)})", "DBEAFE", row)
        for att in agape_members:
            civ = "S∴" if att.member.civility == "S" else "F∴"
            row = _data_row(ws, [
                att.member.last_name,
                f"{civ} {att.member.first_name}",
                "Frère / Sœur",
                "Loge Socrate",
                att.agape_guests if att.agape_guests else "—",
                "—",
            ], C_FRERE, row)

    # Visiteurs
    agape_visitors = sorted([mv for mv in meeting.meeting_visitors
                             if mv.agape and mv.status.value == "CONFIRMED"],
                            key=lambda mv: mv.visitor.last_name)
    if agape_visitors:
        row = _section_row(ws, f"  Maçons passants ({len(agape_visitors)})", "BFDBFE", row)
        for mv in agape_visitors:
            civ = "S∴" if mv.visitor.civility == "S" else "F∴"
            row = _data_row(ws, [
                mv.visitor.last_name,
                f"{civ} {mv.visitor.first_name}",
                "Maçon passant",
                mv.visitor.lodge_name or "—",
                mv.agape_guests if mv.agape_guests else "—",
                "—",
            ], C_VISITEUR, row)

    # Invités profanes
    agape_guests = sorted([g for g in meeting.meeting_guests
                           if g.agape and g.status.value == "CONFIRMED"],
                          key=lambda g: g.last_name)
    if agape_guests:
        row = _section_row(ws, f"  Invités profanes ({len(agape_guests)})", "FDE68A", row)
        for g in agape_guests:
            diet_val = g.dietary_restrictions.value if g.dietary_restrictions else "NONE"
            diet_str = diet_labels.get(diet_val, "—")
            fill = C_DIET if diet_val != "NONE" else C_INVITE
            row = _data_row(ws, [
                g.last_name, g.first_name, "Invité profane", "—", "—", diet_str
            ], fill, row)

    # Ligne total
    total_covers = (
        sum(1 + a.agape_guests for a in agape_members)
        + sum(1 + mv.agape_guests for mv in agape_visitors)
        + len(agape_guests)
    )
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "TOTAL COUVERTS À PRÉPARER"
    ws[f"A{row}"].font = Font(bold=True, size=11, color=C_HEADER)
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=C_TOTAL)
    ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws[f"F{row}"] = total_covers
    ws[f"F{row}"].font = Font(bold=True, size=14, color=C_HEADER)
    ws[f"F{row}"].fill = PatternFill("solid", fgColor=C_TOTAL)
    ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = border

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 22

    # ── Feuille 2 : Récapitulatif ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Récapitulatif")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Récapitulatif agapes"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=C_HEADER)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    recap = [
        ("Membres de la loge", len(agape_members),
         sum(1 + a.agape_guests for a in agape_members)),
        ("Maçons passants",    len(agape_visitors),
         sum(1 + mv.agape_guests for mv in agape_visitors)),
        ("Invités profanes",   len(agape_guests), len(agape_guests)),
    ]
    ws2.cell(row=3, column=1, value="Catégorie").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=3, column=2, value="Personnes").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=3, column=3, value="Couverts").font = Font(bold=True, color="FFFFFF")
    for c in range(1, 4):
        ws2.cell(row=3, column=c).fill = PatternFill("solid", fgColor="374151")
        ws2.cell(row=3, column=c).alignment = Alignment(horizontal="center")
        ws2.cell(row=3, column=c).border = border

    fills = [C_FRERE, C_VISITEUR, C_INVITE]
    for i, (label, persons, covers) in enumerate(recap):
        r = 4 + i
        for c, val in enumerate([label, persons, covers], 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.fill = PatternFill("solid", fgColor=fills[i])
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left",
                                       indent=1 if c == 1 else 0)

    # Total
    ws2.cell(row=7, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=7, column=3, value=total_covers).font = Font(bold=True, size=12, color=C_HEADER)
    for c in range(1, 4):
        ws2.cell(row=7, column=c).fill = PatternFill("solid", fgColor=C_TOTAL)
        ws2.cell(row=7, column=c).border = border
        ws2.cell(row=7, column=c).alignment = Alignment(horizontal="center")
    ws2.row_dimensions[7].height = 22

    # Régimes
    diet_counts: dict[str, int] = {}
    for g in agape_guests:
        v = g.dietary_restrictions.value if g.dietary_restrictions else "NONE"
        if v != "NONE":
            diet_counts[v] = diet_counts.get(v, 0) + 1

    if diet_counts:
        ws2.cell(row=9, column=1, value="Régimes alimentaires").font = Font(bold=True)
        ws2.merge_cells("A9:C9")
        ws2["A9"].fill = PatternFill("solid", fgColor="FEF3C7")
        ws2["A9"].border = border
        for i, (code, count) in enumerate(diet_counts.items()):
            r = 10 + i
            ws2.cell(row=r, column=1, value=diet_labels.get(code, code))
            ws2.cell(row=r, column=2, value=count)
            ws2.cell(row=r, column=3, value="personne(s)")
            for c in range(1, 4):
                ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_DIET)
                ws2.cell(row=r, column=c).border = border
                ws2.cell(row=r, column=c).alignment = Alignment(
                    horizontal="center" if c > 1 else "left", indent=1 if c == 1 else 0
                )

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"agapes_{meeting.meeting_date.strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Page Tracé (Secrétaire) ──────────────────────────────────────────────────

@router.get("/{meeting_id}/trace", response_class=HTMLResponse)
async def meeting_trace(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    if not (can_manage_meeting(member) or user.is_admin):
        raise HTTPException(status_code=403, detail="Accès réservé à la Secrétaire et aux officiers")

    result = await db.execute(
        select(Meeting)
        .options(
            selectinload(Meeting.attendances).selectinload(Attendance.member),
            selectinload(Meeting.meeting_visitors).selectinload(MeetingVisitor.visitor),
            selectinload(Meeting.degrees),
        )
        .where(Meeting.id == meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    # Offices pour afficher les fonctions des membres
    offices_r = await db.execute(
        select(LodgeOffice).where(LodgeOffice.member_id.isnot(None))
    )
    offices = offices_r.scalars().all()
    member_office: dict[int, str] = {o.member_id: o.label for o in offices}

    # Substituts désignés pour cette tenue : remplacent ou complètent member_office
    meeting_subs = (await db.execute(
        select(MeetingOffice).where(MeetingOffice.meeting_id == meeting_id)
    )).scalars().all()
    for ms in meeting_subs:
        if not ms.substitute_member_id:
            continue
        # Le substitut hérite du label de cet office pour cette tenue
        # (note: si le titulaire est aussi présent, le label reste sur le titulaire ;
        # ici on l'écrase pour le substitut s'il n'a pas déjà un autre office)
        if ms.substitute_member_id not in member_office:
            member_office[ms.substitute_member_id] = ms.office_label + " *"

    # Lodge infos
    lodge_r = await db.execute(select(LodgeSettings).limit(1))
    lodge = lodge_r.scalar_one_or_none()

    present  = sorted([a for a in meeting.attendances if a.status == AttendanceStatus.PRESENT],
                      key=lambda a: a.member.last_name)
    excused  = sorted([a for a in meeting.attendances if a.status == AttendanceStatus.EXCUSED],
                      key=lambda a: a.member.last_name)
    absent   = sorted([a for a in meeting.attendances if a.status == AttendanceStatus.ABSENT],
                      key=lambda a: a.member.last_name)
    visitors = sorted([mv for mv in meeting.meeting_visitors
                       if mv.status.value == "CONFIRMED"],
                      key=lambda mv: mv.visitor.last_name)

    # ── Date maçonnique ────────────────────────────────────────────────────
    d = meeting.meeting_date
    masonic_year  = d.year + 4000
    masonic_month = d.month - 2 if d.month >= 3 else d.month + 10
    day_suffix    = "er" if d.day == 1 else "ème"
    month_suffix  = "er" if masonic_month == 1 else "ème"

    # VM de la loge (office label = VM)
    vm_office = next((o for o in offices if o.label == "VM"), None)
    vm_name = ""
    if vm_office and vm_office.member_id:
        for att in present:
            if att.member_id == vm_office.member_id:
                vm_name = f"{att.member.first_name} {att.member.last_name}"
                break

    can_edit = can_manage_meeting(member) or user.is_admin

    # Audit consultation (si activé via /admin/confidentiality)
    try:
        from app.services.confidentiality import maybe_audit_view
        await maybe_audit_view(
            db, actor_id=member.id,
            resource_type="meeting_trace", resource_id=meeting.id,
            target_label=f"Tracé tenue du {meeting.meeting_date.strftime('%d/%m/%Y')}",
            request=request,
        )
    except Exception:
        pass

    return templates.TemplateResponse(request, "pages/meetings/trace.html", {
        "current_member": member,
        "current_user": user,
        "meeting": meeting,
        "lodge": lodge,
        "present": present,
        "excused": excused,
        "absent": absent,
        "visitors": visitors,
        "member_office": member_office,
        "type_label": _type_label,
        "grade_label": _grade_label,
        "masonic_year": masonic_year,
        "masonic_month": masonic_month,
        "day_suffix": day_suffix,
        "month_suffix": month_suffix,
        "vm_name": vm_name,
        "can_edit": can_edit,
    })


# ── Sauvegarde du corps narratif du tracé ────────────────────────────────────

@router.post("/{meeting_id}/trace/save")
async def trace_save(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    compte_rendu_html: str = Form(""),
):
    from fastapi.responses import JSONResponse
    user, member = ctx
    if not (can_manage_meeting(member) or user.is_admin):
        raise HTTPException(status_code=403)

    result = await db.execute(select(Meeting).where(Meeting.id == meeting_id))
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    meeting.compte_rendu_html = compte_rendu_html or None
    await db.commit()
    return JSONResponse({"ok": True})


# ── Page Banquet (Maître des banquets) ───────────────────────────────────────

@router.get("/{meeting_id}/banquet", response_class=HTMLResponse)
async def meeting_banquet(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    can_banquet = (
        user.is_admin
        or can_manage_meeting(member)
        or member.lodge_function == LodgeFunction.MAITRE_BANQUETS
    )
    if not can_banquet:
        raise HTTPException(status_code=403, detail="Accès réservé au Maître des Banquets")

    result = await db.execute(
        select(Meeting)
        .options(
            selectinload(Meeting.attendances).selectinload(Attendance.member),
            selectinload(Meeting.meeting_visitors).selectinload(MeetingVisitor.visitor),
            selectinload(Meeting.meeting_guests),
            selectinload(Meeting.waitlist),
        )
        .where(Meeting.id == meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    agape_members  = sorted([a for a in meeting.attendances if a.agape],
                            key=lambda a: a.member.last_name)
    agape_visitors = sorted([mv for mv in meeting.meeting_visitors
                             if mv.agape and mv.status.value == "CONFIRMED"],
                            key=lambda mv: mv.visitor.last_name)
    agape_guests   = sorted([g for g in meeting.meeting_guests
                             if g.agape and g.status.value == "CONFIRMED"],
                            key=lambda g: g.last_name)

    total_covers = (
        sum(1 + a.agape_guests for a in agape_members)
        + sum(1 + mv.agape_guests for mv in agape_visitors)
        + len(agape_guests)
    )

    # ── Substituts désignés pour cette tenue (depuis Conseil d'officiers) ──
    subs_qs = await db.execute(
        select(MeetingOffice).where(
            MeetingOffice.meeting_id == meeting_id,
            MeetingOffice.substitute_member_id.isnot(None),
        )
    )
    subs_by_office = {s.office_label: s for s in subs_qs.scalars().all()}
    sub_member_ids = {s.substitute_member_id for s in subs_by_office.values()}
    sub_members_cache: dict[int, Member] = {}
    if sub_member_ids:
        sm = await db.execute(select(Member).where(Member.id.in_(sub_member_ids)))
        for m in sm.scalars().all():
            sub_members_cache[m.id] = m

    # Résumé régimes
    diet_counts: dict[str, int] = {}
    for g in agape_guests:
        v = g.dietary_restrictions.value if g.dietary_restrictions else "NONE"
        if v != "NONE":
            diet_counts[v] = diet_counts.get(v, 0) + 1

    diet_labels = {
        "VEGETARIAN": "Végétarien",
        "VEGAN": "Vegan",
        "NO_PORK": "Sans porc",
        "OTHER": "Régime spécial",
    }

    return templates.TemplateResponse(request, "pages/meetings/banquet.html", {
        "current_member": member,
        "current_user": user,
        "meeting": meeting,
        "agape_members": agape_members,
        "agape_visitors": agape_visitors,
        "agape_guests": agape_guests,
        "subs_by_office": subs_by_office,
        "sub_members_cache": sub_members_cache,
        "total_covers": total_covers,
        "diet_counts": diet_counts,
        "diet_labels": diet_labels,
        "type_label": _type_label,
    })


# ── Tableau "Mes agapes" pour le Maître des Banquets ──────────────────────────

@router.get("/me/agapes", response_class=HTMLResponse)
async def my_agapes(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Vue dédiée Maître des Banquets : prochaines tenues avec agapes."""
    user, member = ctx
    is_mdb = (
        user.is_admin
        or can_manage_meeting(member)
        or (member and member.lodge_function == LodgeFunction.MAITRE_BANQUETS)
    )
    if not is_mdb:
        raise HTTPException(403, "Accès réservé au Maître des Banquets")

    today = date.today()
    # Prochaines tenues avec agapes activées + counts
    upcoming = (await db.execute(
        select(Meeting).where(
            Meeting.meeting_date >= today,
            Meeting.agape_enabled == True,  # noqa: E712
        ).order_by(Meeting.meeting_date.asc())
    )).scalars().all()

    rows = []
    for m in upcoming:
        covers = await _count_agapes(db, m.id)
        rows.append({"meeting": m, "covers": covers})

    return templates.TemplateResponse(request, "pages/meetings/my_agapes.html", {
        "current_user": user, "current_member": member,
        "rows": rows, "today": today,
    })


# ── Conseil d'officiers (VM) ──────────────────────────────────────────────────

def _can_manage_officiers(user, member) -> bool:
    """VM, Secrétaire, 1er/2e Surveillants ou admin peuvent gérer le conseil."""
    if user.is_admin:
        return True
    if not member or not member.lodge_function:
        return False
    return member.lodge_function in (
        LodgeFunction.VM, LodgeFunction.SECRETAIRE,
        LodgeFunction.PREMIER_S, LodgeFunction.SECOND_S,
    )


@router.get("/me/officiers")
async def my_officiers(
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Redirige vers le Conseil d'officiers de la prochaine tenue (VM/Surveillants)."""
    user, member = ctx
    today = date.today()
    r = await db.execute(
        select(Meeting).where(Meeting.meeting_date >= today)
        .order_by(Meeting.meeting_date.asc())
        .limit(1)
    )
    m = r.scalar_one_or_none()
    if not m:
        return RedirectResponse(url="/meetings/", status_code=303)
    return RedirectResponse(url=f"/meetings/{m.id}/officiers", status_code=303)


@router.get("/{meeting_id}/officiers", response_class=HTMLResponse)
async def meeting_officiers(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Vue VM : statut de chaque office pour cette tenue + sélecteur de remplaçants."""
    user, member = ctx
    if not _can_manage_officiers(user, member):
        raise HTTPException(403, "Accès réservé au VM, aux Surveillants et au Secrétaire")

    meeting = (await db.execute(
        select(Meeting)
        .options(selectinload(Meeting.attendances).selectinload(Attendance.member))
        .where(Meeting.id == meeting_id)
    )).scalar_one_or_none()
    if not meeting:
        raise HTTPException(404)

    # Tous les offices configurés
    offices = (await db.execute(
        select(LodgeOffice).order_by(LodgeOffice.sort_order, LodgeOffice.label)
    )).scalars().all()

    # Cache titulaires
    holder_ids = {o.member_id for o in offices if o.member_id}
    holders: dict[int, Member] = {}
    if holder_ids:
        mr = await db.execute(select(Member).where(Member.id.in_(holder_ids)))
        for m in mr.scalars().all():
            holders[m.id] = m

    # Statut de présence par membre pour cette tenue
    presence_by_member: dict[int, str] = {}
    for att in meeting.attendances:
        presence_by_member[att.member_id] = att.status.value

    # Remplaçants déjà désignés (table meeting_offices)
    subs = (await db.execute(
        select(MeetingOffice).where(MeetingOffice.meeting_id == meeting_id)
    )).scalars().all()
    sub_by_label: dict[str, MeetingOffice] = {s.office_label: s for s in subs}

    # Cache substituts
    sub_member_ids = {s.substitute_member_id for s in subs if s.substitute_member_id}
    if sub_member_ids:
        sr = await db.execute(select(Member).where(Member.id.in_(sub_member_ids)))
        for m in sr.scalars().all():
            holders[m.id] = m

    # Membres présents/excusés/confirmés (pour le sélecteur de remplaçants)
    present_members = sorted(
        [att.member for att in meeting.attendances
         if att.status == AttendanceStatus.PRESENT and att.member],
        key=lambda x: (x.last_name or "", x.first_name or ""),
    )

    # Tous les membres actifs (pour permettre de désigner même un non-inscrit)
    from app.models.identity import MemberStatus
    all_active = (await db.execute(
        select(Member).where(Member.status == MemberStatus.ACTIVE)
        .order_by(Member.last_name, Member.first_name)
    )).scalars().all()

    # Construction des rangées pour le template
    rows = []
    for o in offices:
        holder = holders.get(o.member_id) if o.member_id else None
        holder_status = presence_by_member.get(o.member_id) if o.member_id else None
        sub_record = sub_by_label.get(o.label)
        sub_member = holders.get(sub_record.substitute_member_id) if sub_record and sub_record.substitute_member_id else None
        # Est-ce que ça pose problème ? Titulaire absent/excusé/sans réponse
        needs_substitute = (
            not o.member_id  # vacant
            or holder_status in (AttendanceStatus.ABSENT.value, AttendanceStatus.EXCUSED.value)
            or holder_status is None  # pas de réponse
        )
        rows.append({
            "office": o,
            "holder": holder,
            "holder_status": holder_status,
            "needs_substitute": needs_substitute,
            "substitute": sub_member,
            "substitute_id": (sub_record.substitute_member_id if sub_record else None),
            "notes": (sub_record.notes if sub_record else ""),
        })

    return templates.TemplateResponse(request, "pages/meetings/officiers.html", {
        "current_user": user,
        "current_member": member,
        "meeting": meeting,
        "rows": rows,
        "present_members": present_members,
        "all_active": all_active,
        "AttendanceStatus": AttendanceStatus,
    })


@router.get("/{meeting_id}/officiers/export.pdf")
async def meeting_officiers_pdf(
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Génère un PDF du conseil d'officiers via ReportLab."""
    from fastapi.responses import StreamingResponse
    import io
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except Exception:
        raise HTTPException(500, "ReportLab manquant")

    user, member = ctx
    if not _can_manage_officiers(user, member):
        raise HTTPException(403)

    meeting = (await db.execute(
        select(Meeting).options(selectinload(Meeting.attendances).selectinload(Attendance.member))
        .where(Meeting.id == meeting_id)
    )).scalar_one_or_none()
    if not meeting:
        raise HTTPException(404)

    offices = (await db.execute(select(LodgeOffice).order_by(LodgeOffice.sort_order, LodgeOffice.label))).scalars().all()
    holder_ids = {o.member_id for o in offices if o.member_id}
    holders: dict[int, Member] = {}
    if holder_ids:
        for m in (await db.execute(select(Member).where(Member.id.in_(holder_ids)))).scalars().all():
            holders[m.id] = m
    presence = {att.member_id: att.status.value for att in meeting.attendances}
    subs = (await db.execute(select(MeetingOffice).where(MeetingOffice.meeting_id == meeting_id))).scalars().all()
    sub_map = {s.office_label: s for s in subs}
    sub_ids = {s.substitute_member_id for s in subs if s.substitute_member_id}
    if sub_ids:
        for m in (await db.execute(select(Member).where(Member.id.in_(sub_ids)))).scalars().all():
            holders[m.id] = m
    lodge_s = (await db.execute(select(LodgeSettings).limit(1))).scalar_one_or_none()
    lodge_name = lodge_s.name if lodge_s else "Loge Socrate"

    # Build PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2.5*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    teal = colors.HexColor("#2c7a7b")
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], textColor=teal, fontSize=16)
    h2 = ParagraphStyle("H2", parent=styles["Normal"], textColor=teal, fontSize=9, fontName="Helvetica-Bold")
    normal = ParagraphStyle("N", parent=styles["Normal"], fontSize=9)

    STATUS_LABEL = {"PRESENT": "✓ Présent", "EXCUSED": "⚠ Excusé", "ABSENT": "✗ Absent"}
    STATUS_COLOR = {"PRESENT": colors.HexColor("#047857"), "EXCUSED": colors.HexColor("#b45309"), "ABSENT": colors.HexColor("#be123c")}

    elems = [
        Paragraph(lodge_name, h1),
        Paragraph(f"Conseil d'officiers — {meeting.title or meeting.type.value} du {meeting.meeting_date.strftime('%d/%m/%Y')}", normal),
        Spacer(1, 0.5*cm),
    ]
    header_row = [Paragraph(h, h2) for h in ["Office", "Titulaire", "Présence", "Remplaçant", "Notes"]]
    data = [header_row]
    for o in offices:
        holder = holders.get(o.member_id) if o.member_id else None
        st = presence.get(o.member_id, "") if o.member_id else ""
        sub_record = sub_map.get(o.label)
        sub = holders.get(sub_record.substitute_member_id) if sub_record and sub_record.substitute_member_id else None
        holder_name = f"{holder.first_name} {holder.last_name}" if holder else "Vacant"
        sub_name = f"{sub.first_name} {sub.last_name}" if sub else "—"
        st_label = STATUS_LABEL.get(st, "—")
        st_color = STATUS_COLOR.get(st, colors.HexColor("#6b7280"))
        data.append([
            Paragraph(o.label, normal),
            Paragraph(holder_name, normal),
            Paragraph(f'<font color="{st_color.hexval()}">{st_label}</font>', normal) if st else Paragraph("—", normal),
            Paragraph(sub_name, ParagraphStyle("sub", parent=normal, textColor=teal) if sub else normal),
            Paragraph(sub_record.notes or "" if sub_record else "", normal),
        ])

    col_w = [3.5*cm, 4*cm, 3*cm, 4*cm, 3*cm]
    tbl = Table(data, colWidths=col_w)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#e6f4f1")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#cccccc")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f7fafa")]),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="officiers-{meeting_id}.pdf"'},
    )


@router.get("/{meeting_id}/officiers/print", response_class=HTMLResponse)
async def meeting_officiers_print(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Vue imprimable du conseil d'officiers — A4 portrait."""
    user, member = ctx
    if not _can_manage_officiers(user, member):
        raise HTTPException(403)

    meeting = (await db.execute(
        select(Meeting)
        .options(selectinload(Meeting.attendances).selectinload(Attendance.member))
        .where(Meeting.id == meeting_id)
    )).scalar_one_or_none()
    if not meeting:
        raise HTTPException(404)

    offices = (await db.execute(
        select(LodgeOffice).order_by(LodgeOffice.sort_order, LodgeOffice.label)
    )).scalars().all()

    holder_ids = {o.member_id for o in offices if o.member_id}
    holders: dict[int, Member] = {}
    if holder_ids:
        mr = await db.execute(select(Member).where(Member.id.in_(holder_ids)))
        for m in mr.scalars().all():
            holders[m.id] = m

    presence_by_member = {att.member_id: att.status.value for att in meeting.attendances}

    subs = (await db.execute(
        select(MeetingOffice).where(MeetingOffice.meeting_id == meeting_id)
    )).scalars().all()
    sub_by_label = {s.office_label: s for s in subs}

    sub_member_ids = {s.substitute_member_id for s in subs if s.substitute_member_id}
    if sub_member_ids:
        sr = await db.execute(select(Member).where(Member.id.in_(sub_member_ids)))
        for m in sr.scalars().all():
            holders[m.id] = m

    rows = []
    for o in offices:
        sub_record = sub_by_label.get(o.label)
        sub_member = holders.get(sub_record.substitute_member_id) if sub_record and sub_record.substitute_member_id else None
        rows.append({
            "office": o,
            "holder": holders.get(o.member_id) if o.member_id else None,
            "holder_status": presence_by_member.get(o.member_id) if o.member_id else None,
            "substitute": sub_member,
            "notes": sub_record.notes if sub_record else "",
        })

    lodge = (await db.execute(select(LodgeSettings).limit(1))).scalar_one_or_none()

    return templates.TemplateResponse(request, "pages/meetings/officiers_print.html", {
        "current_user": user,
        "current_member": member,
        "meeting": meeting,
        "rows": rows,
        "lodge": lodge,
    })


@router.post("/{meeting_id}/officiers/save")
async def meeting_officiers_save(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Sauvegarde des remplaçants. Inputs : `sub[label]` et `notes[label]`."""
    user, member = ctx
    if not _can_manage_officiers(user, member):
        raise HTTPException(403)

    meeting = (await db.execute(select(Meeting).where(Meeting.id == meeting_id))).scalar_one_or_none()
    if not meeting:
        raise HTTPException(404)

    form = await request.form()

    # Récupère les enregistrements existants
    existing_rows = (await db.execute(
        select(MeetingOffice).where(MeetingOffice.meeting_id == meeting_id)
    )).scalars().all()
    existing_by_label: dict[str, MeetingOffice] = {r.office_label: r for r in existing_rows}

    # Parse les champs
    subs_input: dict[str, dict] = {}
    for k, v in form.multi_items():
        if k.startswith("sub[") and k.endswith("]"):
            label = k[4:-1]
            sub_id = None
            if v and v.strip().isdigit():
                sub_id = int(v.strip())
            subs_input.setdefault(label, {})["sub_id"] = sub_id
        elif k.startswith("notes[") and k.endswith("]"):
            label = k[6:-1]
            subs_input.setdefault(label, {})["notes"] = (v or "").strip()[:300]

    for label, data in subs_input.items():
        sub_id = data.get("sub_id")
        notes = data.get("notes", "")
        row = existing_by_label.get(label)
        # Si rien renseigné, supprimer la ligne existante
        if not sub_id and not notes:
            if row:
                await db.delete(row)
            continue
        if row:
            row.substitute_member_id = sub_id
            row.notes = notes or None
        else:
            db.add(MeetingOffice(
                meeting_id=meeting_id,
                office_label=label,
                substitute_member_id=sub_id,
                notes=notes or None,
            ))
    await db.commit()
    return RedirectResponse(url=f"/meetings/{meeting_id}/officiers?_saved=1", status_code=303)


# ── Formulaire nouvelle tenue ─────────────────────────────────────────────────

@router.get("/new/form", response_class=HTMLResponse)
async def meeting_new_form(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    if not (can_manage_meeting(member) or user.is_admin):
        raise HTTPException(status_code=403, detail="Accès refusé")

    # Année courante
    year_result = await db.execute(select(MasonicYear).where(MasonicYear.is_current == True))
    current_year = year_result.scalar_one_or_none()

    return templates.TemplateResponse(request, "pages/meetings/form.html", {
        "current_member": member,
        "current_user": user,
        "meeting": None,
        "current_year": current_year,
        "MeetingType": MeetingType,
        "MeetingGrade": MeetingGrade,
        "type_label": _type_label,
        "grade_label": _grade_label,
        "errors": {},
        "is_new": True,
        "form_action": "/meetings/new/form",
        "degree_labels": {
            "APPRENTI": "1er degré — Apprentis",
            "COMPAGNON": "2e degré — Compagnons",
            "MAITRE": "3e degré — Maîtres",
            "ALL": "Toutes loges réunies",
        },
    })


@router.post("/new/form", response_class=HTMLResponse)
async def meeting_create(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    meeting_date:    str = Form(...),
    meeting_time:    str = Form("20:30"),
    meeting_type:    str = Form("BLANCHE"),
    meeting_grade:   str = Form("MAITRE"),
    meeting_number:  str = Form(""),
    title:           str = Form(""),
    theme:           str = Form(""),
    location:        str = Form(""),
    address:         str = Form(""),
    agenda_html:     str = Form(""),
    agape_enabled:   str = Form(""),
    agape_capacity:  str = Form(""),
    agape_location:  str = Form(""),
    visio_url:       str = Form(""),
    # Multi-degrés : liste de degrés séparés par virgule + descriptions
    degrees_grades:  str = Form(""),  # ex: "APPRENTI,COMPAGNON,MAITRE"
    degrees_descs:   str = Form(""),  # ex: "Ouverture,Passage,Travaux"
):
    user, member = ctx
    if not (can_manage_meeting(member) or user.is_admin):
        raise HTTPException(status_code=403, detail="Accès refusé")

    # Récupérer l'année maçonnique correspondante
    d = date.fromisoformat(meeting_date)
    year_result = await db.execute(
        select(MasonicYear).where(
            MasonicYear.start_date <= d,
            MasonicYear.end_date >= d,
        )
    )
    masonic_year = year_result.scalar_one_or_none()
    if not masonic_year:
        # Fallback : année courante
        year_result = await db.execute(select(MasonicYear).where(MasonicYear.is_current == True))
        masonic_year = year_result.scalar_one_or_none()

    # Date de clôture des inscriptions = veille J-1 à 8h du matin
    reg_closes = datetime(d.year, d.month, d.day, 8, 0, 0) - timedelta(days=1)

    # Auto-générer lien visio si configuré et aucun lien manuel fourni
    final_visio_url = visio_url.strip() or None
    if not final_visio_url:
        ls_r = await db.execute(select(LodgeSettings).limit(1))
        lodge_cfg = ls_r.scalar_one_or_none()
        if lodge_cfg and lodge_cfg.visio_provider and lodge_cfg.visio_server_url:
            prefix = lodge_cfg.visio_room_prefix or "loge"
            room   = f"{prefix}-{d.strftime('%Y%m%d')}"
            base   = lodge_cfg.visio_server_url.rstrip("/")
            final_visio_url = f"{base}/{room}"

    new_meeting = Meeting(
        masonic_year_id=masonic_year.id if masonic_year else 1,
        meeting_date=d,
        meeting_time=meeting_time or "20:30",
        type=MeetingType(meeting_type),
        grade=MeetingGrade(meeting_grade),
        meeting_number=int(meeting_number) if meeting_number.strip().isdigit() else None,
        title=title or None,
        theme=theme or None,
        location=location or None,
        address=address or None,
        agenda_html=agenda_html or None,
        agape_enabled=bool(agape_enabled),
        agape_capacity=int(agape_capacity) if agape_capacity else None,
        agape_location=agape_location or None,
        visio_url=final_visio_url,
        registration_closes_at=reg_closes,
        created_by_id=member.id,
    )
    db.add(new_meeting)
    await db.flush()  # obtenir l'ID

    # Enregistrer la séquence des degrés si multi-degrés
    if degrees_grades.strip():
        grades_list = [g.strip() for g in degrees_grades.split(",") if g.strip()]
        descs_list  = [d.strip() for d in degrees_descs.split("|") if True]  # séparateur |
        for i, grade_str in enumerate(grades_list):
            try:
                g = MeetingGrade(grade_str)
            except ValueError:
                continue
            desc = descs_list[i] if i < len(descs_list) else ""
            deg = MeetingDegree(
                meeting_id=new_meeting.id,
                order_position=i + 1,
                grade=g,
                description=desc or None,
            )
            db.add(deg)

    await db.commit()

    # ── Push notification aux membres concernés (grade >= meeting.grade) ─────
    try:
        from app.models.identity import MemberStatus, MasonicGrade
        from app.services.push import send_push_broadcast
        GRADE_LEVEL = {MasonicGrade.APPRENTI: 1, MasonicGrade.COMPAGNON: 2, MasonicGrade.MAITRE: 3}
        meeting_level = {"APPRENTI": 1, "COMPAGNON": 2, "MAITRE": 3, "BLANCHE": 1}.get(new_meeting.grade.value, 1)
        r = await db.execute(select(Member).where(Member.status == MemberStatus.ACTIVE))
        eligible_ids = [
            m.id for m in r.scalars().all()
            if m.id != member.id and GRADE_LEVEL.get(m.masonic_grade, 0) >= meeting_level
        ]
        date_str = d.strftime("%d/%m/%Y")
        title_push = f"📅 Nouvelle tenue le {date_str}"
        body_push = (new_meeting.title or new_meeting.theme or f"Tenue {new_meeting.grade.value.lower()}")[:140]
        await send_push_broadcast(db, eligible_ids, title_push, body_push, f"/meetings/{new_meeting.id}")
    except Exception:
        pass

    return RedirectResponse(url=f"/meetings/{new_meeting.id}", status_code=302)


# ── Édition d'une tenue ──────────────────────────────────────────────────────

@router.get("/{meeting_id}/edit", response_class=HTMLResponse)
async def meeting_edit_form(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    if not (can_manage_meeting(member) or user.is_admin):
        raise HTTPException(status_code=403, detail="Accès refusé")

    result = await db.execute(
        select(Meeting)
        .options(selectinload(Meeting.degrees))
        .where(Meeting.id == meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404, detail="Tenue introuvable")

    year_result = await db.execute(select(MasonicYear).where(MasonicYear.is_current == True))
    current_year = year_result.scalar_one_or_none()

    return templates.TemplateResponse(request, "pages/meetings/form.html", {
        "current_member": member,
        "current_user": user,
        "meeting": meeting,
        "current_year": current_year,
        "MeetingType": MeetingType,
        "MeetingGrade": MeetingGrade,
        "type_label": _type_label,
        "grade_label": _grade_label,
        "errors": {},
        "is_new": False,
        "form_action": f"/meetings/{meeting_id}/edit",
        "degree_labels": {
            "APPRENTI": "1er degré — Apprentis",
            "COMPAGNON": "2e degré — Compagnons",
            "MAITRE": "3e degré — Maîtres",
            "ALL": "Toutes loges réunies",
        },
    })


@router.post("/{meeting_id}/edit", response_class=HTMLResponse)
async def meeting_edit_save(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    meeting_date:    str = Form(...),
    meeting_time:    str = Form("20:30"),
    meeting_type:    str = Form("BLANCHE"),
    meeting_grade:   str = Form("MAITRE"),
    meeting_number:  str = Form(""),
    title:           str = Form(""),
    theme:           str = Form(""),
    location:        str = Form(""),
    address:         str = Form(""),
    agenda_html:     str = Form(""),
    agape_enabled:   str = Form(""),
    agape_capacity:  str = Form(""),
    agape_location:  str = Form(""),
    visio_url:       str = Form(""),
    degrees_grades:  str = Form(""),
    degrees_descs:   str = Form(""),
):
    user, member = ctx
    if not (can_manage_meeting(member) or user.is_admin):
        raise HTTPException(status_code=403, detail="Accès refusé")

    result = await db.execute(
        select(Meeting)
        .options(selectinload(Meeting.degrees))
        .where(Meeting.id == meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    d = date.fromisoformat(meeting_date)

    meeting.meeting_date    = d
    meeting.meeting_time    = meeting_time or "20:30"
    meeting.type            = MeetingType(meeting_type)
    meeting.grade           = MeetingGrade(meeting_grade)
    meeting.meeting_number  = int(meeting_number) if meeting_number.strip().isdigit() else None
    meeting.title           = title or None
    meeting.theme           = theme or None
    meeting.location        = location or None
    meeting.address         = address or None
    meeting.agenda_html     = agenda_html or None
    meeting.agape_enabled   = bool(agape_enabled)
    meeting.agape_capacity  = int(agape_capacity) if agape_capacity else None
    meeting.agape_location  = agape_location or None
    meeting.visio_url       = visio_url or None

    # Reconstruire la séquence des degrés
    for deg in list(meeting.degrees):
        await db.delete(deg)
    await db.flush()

    if degrees_grades.strip():
        grades_list = [g.strip() for g in degrees_grades.split(",") if g.strip()]
        descs_list  = [d.strip() for d in degrees_descs.split("|")]
        for i, grade_str in enumerate(grades_list):
            try:
                g = MeetingGrade(grade_str)
            except ValueError:
                continue
            desc = descs_list[i] if i < len(descs_list) else ""
            db.add(MeetingDegree(
                meeting_id=meeting.id,
                order_position=i + 1,
                grade=g,
                description=desc or None,
            ))

    await db.commit()
    return RedirectResponse(url=f"/meetings/{meeting_id}", status_code=302)


# ── Suppression d'une tenue ───────────────────────────────────────────────────

@router.post("/{meeting_id}/delete", response_class=HTMLResponse)
async def meeting_delete(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    if not (user.is_admin):
        raise HTTPException(status_code=403, detail="Seul un administrateur peut supprimer une tenue")

    result = await db.execute(select(Meeting).where(Meeting.id == meeting_id))
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    await db.delete(meeting)
    await db.commit()
    return RedirectResponse(url="/meetings/", status_code=302)


# ── Inscription interne (membre connecté) ─────────────────────────────────────

@router.post("/{meeting_id}/register", response_class=HTMLResponse)
async def meeting_register(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    attendance_status: str = Form("PRESENT"),
    agape:             str = Form(""),
    agape_guests:      str = Form("0"),
    excuse_reason:     str = Form(""),
):
    user, member = ctx

    result = await db.execute(select(Meeting).where(Meeting.id == meeting_id))
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    # Vérifier si déjà inscrit
    existing = await db.execute(
        select(Attendance).where(
            Attendance.meeting_id == meeting_id,
            Attendance.member_id == member.id,
        )
    )
    att = existing.scalar_one_or_none()

    if att:
        att.status = AttendanceStatus(attendance_status)
        att.agape = bool(agape)
        att.agape_guests = int(agape_guests) if agape_guests else 0
        att.excuse_reason = excuse_reason or None
    else:
        att = Attendance(
            meeting_id=meeting_id,
            member_id=member.id,
            status=AttendanceStatus(attendance_status),
            agape=bool(agape),
            agape_guests=int(agape_guests) if agape_guests else 0,
            excuse_reason=excuse_reason or None,
        )
        db.add(att)

    await db.commit()
    return RedirectResponse(url=f"/meetings/{meeting_id}", status_code=302)


# ── Inscription publique (lien depuis le programme PDF) ───────────────────────

@router.get("/public/{token}", response_class=HTMLResponse)
async def public_register_page(
    request: Request,
    token: str,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Page publique d'inscription — accessible sans login."""
    result = await db.execute(select(Meeting).where(Meeting.token == token))
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404, detail="Lien d'inscription invalide")

    # Fermeture auto : tenue passée OU date de clôture dépassée
    now = datetime.now()
    should_close = (
        meeting.meeting_date < date.today()
        or (meeting.registration_closes_at is not None and meeting.registration_closes_at <= now)
    )
    if should_close and meeting.registration_open:
        meeting.registration_open = False
        await db.commit()

    if not meeting.registration_open:
        return templates.TemplateResponse(request, "pages/meetings/register_closed.html", {
            "meeting": meeting,
            "type_label": _type_label,
        })

    from app.models.lodge import LodgeSettings as _LS
    r2 = await db.execute(select(_LS).limit(1))
    lodge = r2.scalar_one_or_none()

    # Compteur agapes actuel
    agape_count = await _count_agapes(db, meeting.id)

    # Liste membres actifs (pour dropdown PIN)
    members_r = await db.execute(
        select(Member)
        .where(Member.status == "ACTIVE")
        .order_by(Member.last_name, Member.first_name)
    )
    active_members = members_r.scalars().all()

    from sqlalchemy import distinct as sql_distinct
    lodges_r = await db.execute(select(sql_distinct(Visitor.lodge_name)).where(Visitor.lodge_name.isnot(None), Visitor.lodge_name != ""))
    orients_r = await db.execute(select(sql_distinct(Visitor.orient_city)).where(Visitor.orient_city.isnot(None), Visitor.orient_city != ""))
    obeds_r = await db.execute(select(sql_distinct(Visitor.obedience)).where(Visitor.obedience.isnot(None), Visitor.obedience != ""))

    return templates.TemplateResponse(request, "pages/meetings/register_public.html", {
        "meeting": meeting,
        "token": token,
        "type_label": _type_label,
        "DietaryRestriction": DietaryRestriction,
        "lodge": lodge,
        "agape_count": agape_count,
        "active_members": active_members,
        "known_lodges": sorted(lodges_r.scalars().all()),
        "known_orients": sorted(orients_r.scalars().all()),
        "known_obediences": sorted(obeds_r.scalars().all()),
    })


@router.post("/public/{token}", response_class=HTMLResponse)
async def public_register_submit(
    request: Request,
    token: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    visitor_type:      str = Form("visitor"),  # "member" | "visitor"
    # Inscription membre (PIN)
    member_id:         str = Form(""),
    pin_code:          str = Form(""),
    # Inscription visiteur
    last_name:         str = Form(""),
    first_name:        str = Form(""),
    email:             str = Form(""),
    civility:          str = Form("F"),
    lodge_name:        str = Form(""),
    orient_city:       str = Form(""),
    obedience:         str = Form(""),
    masonic_grade_str: str = Form(""),
    is_vm:             str = Form(""),
    comment:           str = Form(""),
    agape:             str = Form(""),
    agape_guests:      str = Form("0"),
    phone:             str = Form(""),
):
    """Traitement de l'inscription publique."""
    from app.dependencies import verify_password

    result = await db.execute(select(Meeting).where(Meeting.token == token))
    meeting = result.scalar_one_or_none()
    if not meeting or not meeting.registration_open:
        raise HTTPException(status_code=404)

    agape_bool = bool(agape)
    agape_guests_int = max(0, int(agape_guests) if agape_guests.isdigit() else 0)

    # ── Vérifier capacité agapes ───────────────────────────────────────────
    places_needed = (1 + agape_guests_int) if agape_bool else 0
    if meeting.agape_capacity and agape_bool:
        current = await _count_agapes(db, meeting.id)
        remaining = meeting.agape_capacity - current
        if places_needed > remaining:
            # Liste d'attente
            pos_r = await db.execute(
                select(sql_func.count()).select_from(MeetingWaitlist)
                .where(MeetingWaitlist.meeting_id == meeting.id)
            )
            position = (pos_r.scalar() or 0) + 1
            db.add(MeetingWaitlist(
                meeting_id=meeting.id,
                external_name=f"{first_name} {last_name}".strip() or f"Membre #{member_id}",
                external_email=email or None,
                position=position,
            ))
            await db.commit()
            return templates.TemplateResponse(request, "pages/meetings/register_waitlist.html", {
                "meeting": meeting, "position": position, "type_label": _type_label,
            })

    # ── Cas 1 : Membre de la loge (PIN) ───────────────────────────────────
    if visitor_type == "member":
        error = None
        target_member = None

        if not member_id.isdigit():
            error = "Veuillez sélectionner votre nom."
        else:
            m_r = await db.execute(select(Member).where(Member.id == int(member_id)))
            target_member = m_r.scalar_one_or_none()
            if not target_member:
                error = "Membre introuvable."
            elif not target_member.pin_code_hash:
                error = "Aucun code PIN configuré pour ce compte. Contactez le secrétaire."
            elif not pin_code or not verify_password(pin_code, target_member.pin_code_hash):
                error = "Code PIN incorrect."

        if error:
            from app.models.lodge import LodgeSettings as _LS
            r2 = await db.execute(select(_LS).limit(1))
            lodge = r2.scalar_one_or_none()
            agape_count = await _count_agapes(db, meeting.id)
            members_r = await db.execute(
                select(Member).where(Member.status == "ACTIVE")
                .order_by(Member.last_name, Member.first_name)
            )
            return templates.TemplateResponse(request, "pages/meetings/register_public.html", {
                "meeting": meeting, "token": token, "type_label": _type_label,
                "DietaryRestriction": DietaryRestriction, "lodge": lodge,
                "agape_count": agape_count,
                "active_members": members_r.scalars().all(),
                "pin_error": error,
                "prefill_member_id": member_id,
            }, status_code=422)

        # PIN OK — enregistrer ou mettre à jour la présence
        existing_r = await db.execute(
            select(Attendance).where(
                Attendance.meeting_id == meeting.id,
                Attendance.member_id == target_member.id,
            )
        )
        att = existing_r.scalar_one_or_none()
        if att:
            att.agape = agape_bool
            att.agape_guests = agape_guests_int
            # Ne pas écraser un statut déjà saisi par l'admin
            if att.status == AttendanceStatus.ABSENT:
                att.status = AttendanceStatus.PRESENT
        else:
            att = Attendance(
                meeting_id=meeting.id,
                member_id=target_member.id,
                status=AttendanceStatus.PRESENT,
                agape=agape_bool,
                agape_guests=agape_guests_int,
            )
            db.add(att)

        await db.commit()
        return templates.TemplateResponse(request, "pages/meetings/register_success.html", {
            "meeting": meeting, "visitor_type": "member",
            "first_name": target_member.first_name,
            "last_name": target_member.last_name,
            "agape": agape_bool, "type_label": _type_label,
        })

    # ── Cas 2 : Maçon passant ──────────────────────────────────────────────
    visitor = Visitor(
        civility=civility if civility in ("F", "S") else "F",
        last_name=last_name.strip().upper(),
        first_name=first_name.strip().title(),
        email=email.strip().lower() if email else None,
        lodge_name=lodge_name or None,
        orient_city=orient_city or None,
        obedience=obedience or None,
        masonic_grade=masonic_grade_str or None,
        is_vm=bool(is_vm),
        phone=phone or None,
    )
    db.add(visitor)
    await db.flush()

    db.add(MeetingVisitor(
        meeting_id=meeting.id,
        visitor_id=visitor.id,
        agape=agape_bool,
        agape_guests=agape_guests_int,
        token_used=token,
        comment=comment.strip() or None,
    ))
    await db.commit()

    return templates.TemplateResponse(request, "pages/meetings/register_success.html", {
        "meeting": meeting, "visitor_type": "visitor",
        "first_name": first_name, "last_name": last_name,
        "agape": agape_bool, "type_label": _type_label,
    })


# ── Verrouillage de la tenue ──────────────────────────────────────────────────

@router.post("/{meeting_id}/lock", response_class=HTMLResponse)
async def meeting_lock(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    if not (can_lock_meeting(member) or user.is_admin):
        raise HTTPException(status_code=403)

    result = await db.execute(select(Meeting).where(Meeting.id == meeting_id))
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    meeting.is_locked = not meeting.is_locked
    if meeting.is_locked:
        meeting.locked_by_id = member.id
        meeting.locked_at = datetime.utcnow()
        meeting.registration_open = False
    else:
        meeting.locked_by_id = None
        meeting.locked_at = None

    await db.commit()
    return RedirectResponse(url=f"/meetings/{meeting_id}", status_code=302)
