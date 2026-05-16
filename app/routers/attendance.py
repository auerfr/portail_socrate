"""Router Présences & Assiduité"""
import csv
import io
from datetime import date
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from sqlalchemy import select, func as sql_func
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import require_auth, can_manage_attendance
from app.models.identity import Member, MemberStatus
from app.models.lodge import MasonicYear, LodgeSettings
from app.models.meetings import (
    Meeting, Attendance, AttendanceStatus,
    MeetingVisitor, VisitorStatus, Visitor, MeetingDegree,
)
from app.models.system import TracingSection, TracingSectionType

router = APIRouter(prefix="/attendance", tags=["attendance"])
from app.template_engine import templates


def _require_attendance_mgr(user, member):
    if not (can_manage_attendance(member) or user.is_admin):
        raise HTTPException(status_code=403, detail="Accès réservé au VM, Secrétaire et Surveillants")


# ── Dashboard assiduité ───────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def attendance_dashboard(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    year_id: int = 0,
):
    user, member = ctx
    _require_attendance_mgr(user, member)

    # Seuils assiduité depuis lodge_settings
    ls_r = await db.execute(select(LodgeSettings).limit(1))
    lodge = ls_r.scalar_one_or_none()
    threshold_warn   = (lodge.attendance_threshold_warn   if lodge else None) or 70
    threshold_danger = (lodge.attendance_threshold_danger if lodge else None) or 50

    # Années maçonniques
    years_r = await db.execute(
        select(MasonicYear).order_by(MasonicYear.is_current.desc(), MasonicYear.start_date.desc())
    )
    years = years_r.scalars().all()
    current_year = next((y for y in years if y.is_current), years[0] if years else None)
    selected_year = next((y for y in years if y.id == year_id), current_year)

    year_filter = (Meeting.masonic_year_id == selected_year.id) if selected_year else True

    # ── Tenues passées (pour grille assiduité) ──────────────────────────────
    past_r = await db.execute(
        select(Meeting)
        .where(year_filter, Meeting.meeting_date <= date.today())
        .order_by(Meeting.meeting_date)
    )
    past_meetings = past_r.scalars().all()
    past_ids = [m.id for m in past_meetings]

    # ── Tenues à venir ──────────────────────────────────────────────────────
    upcoming_r = await db.execute(
        select(Meeting)
        .where(year_filter, Meeting.meeting_date > date.today())
        .order_by(Meeting.meeting_date)
    )
    upcoming_meetings = upcoming_r.scalars().all()
    upcoming_ids = [m.id for m in upcoming_meetings]

    # Inscriptions pour les tenues à venir
    upcoming_members_count = {}   # meeting_id → nb membres inscrits PRESENT
    upcoming_visitors_count = {}  # meeting_id → nb passants CONFIRMED
    if upcoming_ids:
        ua_r = await db.execute(
            select(Attendance.meeting_id, sql_func.count().label("n"))
            .where(
                Attendance.meeting_id.in_(upcoming_ids),
                Attendance.status == AttendanceStatus.PRESENT,
            )
            .group_by(Attendance.meeting_id)
        )
        upcoming_members_count = {row.meeting_id: row.n for row in ua_r}

        uv_r = await db.execute(
            select(MeetingVisitor.meeting_id, sql_func.count().label("n"))
            .where(
                MeetingVisitor.meeting_id.in_(upcoming_ids),
                MeetingVisitor.status == VisitorStatus.CONFIRMED,
            )
            .group_by(MeetingVisitor.meeting_id)
        )
        upcoming_visitors_count = {row.meeting_id: row.n for row in uv_r}

    # ── Membres actifs (hors comptes admin techniques) ──────────────────────
    from app.models.identity import User as _User
    admin_member_ids_r = await db.execute(
        select(_User.member_id).where(_User.is_admin == True, _User.member_id.isnot(None))
    )
    _admin_ids = {row[0] for row in admin_member_ids_r}

    members_r = await db.execute(
        select(Member)
        .where(Member.status == MemberStatus.ACTIVE, Member.id.notin_(_admin_ids))
        .order_by(Member.last_name, Member.first_name)
    )
    active_members = members_r.scalars().all()

    # ── Présences membres (tenues passées) ──────────────────────────────────
    stats = {}   # member_id → {PRESENT: n, EXCUSED: n, ABSENT: n}
    grid  = {}   # member_id → {meeting_id → status_value}
    if past_ids:
        att_r = await db.execute(
            select(Attendance).where(Attendance.meeting_id.in_(past_ids))
        )
        for att in att_r.scalars().all():
            s = stats.setdefault(att.member_id, {"PRESENT": 0, "EXCUSED": 0, "ABSENT": 0})
            s[att.status.value] = s.get(att.status.value, 0) + 1
            grid.setdefault(att.member_id, {})[att.meeting_id] = att.status.value

    # ── Tenues applicables par membre (grade + date d'arrivée) ──────────────
    _grade_order = {"APPRENTI": 1, "COMPAGNON": 2, "MAITRE": 3, "ALL": 0}
    member_applicable: dict[int, set[int]] = {}
    for mbr in active_members:
        grade_val = mbr.masonic_grade.value if hasattr(mbr.masonic_grade, "value") else str(mbr.masonic_grade)
        start = mbr.membership_start_date
        applicable: set[int] = set()
        for mtg in past_meetings:
            if start and mtg.meeting_date < start:
                continue
            mtg_grade = mtg.grade.value if hasattr(mtg.grade, "value") else str(mtg.grade)
            if mtg_grade == "ALL" or _grade_order.get(grade_val, 0) >= _grade_order.get(mtg_grade, 0):
                applicable.add(mtg.id)
        member_applicable[mbr.id] = applicable

    # ── KPIs globaux cumulés ────────────────────────────────────────────────
    g_expected = g_present = g_excused = g_absent = 0
    for mbr in active_members:
        applic = member_applicable[mbr.id]
        s = stats.get(mbr.id, {})
        g_expected += len(applic)
        # compter seulement les statuts pour des tenues applicables
        m_grid = grid.get(mbr.id, {})
        for mid in applic:
            v = m_grid.get(mid, "")
            if v == "PRESENT":  g_present += 1
            elif v == "EXCUSED": g_excused += 1
            elif v == "ABSENT":  g_absent  += 1
    g_pct_present = round(g_present * 100 / g_expected) if g_expected else 0
    g_pct_excused = round(g_excused * 100 / g_expected) if g_expected else 0
    g_pct_absent  = round(g_absent  * 100 / g_expected) if g_expected else 0

    # ── Maçons passants confirmés par tenue passée ──────────────────────────
    visitors_per_meeting = {}   # meeting_id → count
    if past_ids:
        mv_r = await db.execute(
            select(MeetingVisitor)
            .where(
                MeetingVisitor.meeting_id.in_(past_ids),
                MeetingVisitor.status == VisitorStatus.CONFIRMED,
            )
        )
        for mv in mv_r.scalars().all():
            visitors_per_meeting[mv.meeting_id] = visitors_per_meeting.get(mv.meeting_id, 0) + 1

    return templates.TemplateResponse(request, "pages/attendance/dashboard.html", {
        "current_member": member,
        "current_user": user,
        "years": years,
        "selected_year": selected_year,
        # passé
        "meetings": past_meetings,
        "total_meetings": len(past_meetings),
        "active_members": active_members,
        "stats": stats,
        "grid": grid,
        "member_applicable": member_applicable,
        "visitors_per_meeting": visitors_per_meeting,
        # à venir
        "upcoming_meetings": upcoming_meetings,
        "upcoming_members_count": upcoming_members_count,
        "upcoming_visitors_count": upcoming_visitors_count,
        # seuils
        "threshold_warn": threshold_warn,
        "threshold_danger": threshold_danger,
        # KPIs globaux
        "g_expected": g_expected,
        "g_present": g_present,
        "g_excused": g_excused,
        "g_absent": g_absent,
        "g_pct_present": g_pct_present,
        "g_pct_excused": g_pct_excused,
        "g_pct_absent": g_pct_absent,
    })


# ── Émargement d'une tenue ────────────────────────────────────────────────────

@router.get("/meeting/{meeting_id}", response_class=HTMLResponse)
async def emargement_page(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    _require_attendance_mgr(user, member)

    result = await db.execute(
        select(Meeting)
        .options(
            selectinload(Meeting.attendances).selectinload(Attendance.member),
            selectinload(Meeting.meeting_visitors).selectinload(MeetingVisitor.visitor),
        )
        .where(Meeting.id == meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    from app.models.identity import User as _User2
    _admin_ids2_r = await db.execute(
        select(_User2.member_id).where(_User2.is_admin == True, _User2.member_id.isnot(None))
    )
    _admin_ids2 = {row[0] for row in _admin_ids2_r}

    members_r = await db.execute(
        select(Member)
        .where(Member.status == MemberStatus.ACTIVE, Member.id.notin_(_admin_ids2))
        .order_by(Member.last_name, Member.first_name)
    )
    all_active = members_r.scalars().all()

    # Filtrer par grade de la tenue
    _grade_order_e = {"APPRENTI": 1, "COMPAGNON": 2, "MAITRE": 3, "ALL": 0}
    mtg_grade_val = meeting.grade.value if hasattr(meeting.grade, "value") else str(meeting.grade)
    if mtg_grade_val == "ALL":
        active_members = all_active
    else:
        active_members = [
            m for m in all_active
            if _grade_order_e.get(
                m.masonic_grade.value if hasattr(m.masonic_grade, "value") else str(m.masonic_grade), 0
            ) >= _grade_order_e.get(mtg_grade_val, 0)
        ]

    att_by_member = {a.member_id: a for a in meeting.attendances}
    visitors = sorted(meeting.meeting_visitors, key=lambda mv: mv.visitor.last_name)

    # Nomenclature connue pour l'autocomplete
    known_lodges_r = await db.execute(
        select(Visitor.lodge_name).where(Visitor.lodge_name.isnot(None)).distinct().order_by(Visitor.lodge_name)
    )
    known_orients_r = await db.execute(
        select(Visitor.orient_city).where(Visitor.orient_city.isnot(None)).distinct().order_by(Visitor.orient_city)
    )
    known_obediences_r = await db.execute(
        select(Visitor.obedience).where(Visitor.obedience.isnot(None)).distinct().order_by(Visitor.obedience)
    )
    known_lodges    = [r[0] for r in known_lodges_r.all()]
    known_orients   = [r[0] for r in known_orients_r.all()]
    known_obediences = [r[0] for r in known_obediences_r.all()]

    return templates.TemplateResponse(request, "pages/attendance/emargement.html", {
        "current_member": member,
        "current_user": user,
        "meeting": meeting,
        "active_members": active_members,
        "att_by_member": att_by_member,
        "AttendanceStatus": AttendanceStatus,
        "visitors": visitors,
        "VisitorStatus": VisitorStatus,
        "known_lodges": known_lodges,
        "known_orients": known_orients,
        "known_obediences": known_obediences,
    })


@router.post("/meeting/{meeting_id}", response_class=HTMLResponse)
async def emargement_save(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    _require_attendance_mgr(user, member)

    result = await db.execute(select(Meeting).where(Meeting.id == meeting_id))
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    form = await request.form()

    # ── Membres (hors admin, filtrés par grade de la tenue) ─────────────────
    from app.models.identity import User as _User3
    _admin_ids3_r = await db.execute(
        select(_User3.member_id).where(_User3.is_admin == True, _User3.member_id.isnot(None))
    )
    _admin_ids3 = {row[0] for row in _admin_ids3_r}
    _grade_order_s = {"APPRENTI": 1, "COMPAGNON": 2, "MAITRE": 3, "ALL": 0}
    _mtg_grade = meeting.grade.value if hasattr(meeting.grade, "value") else str(meeting.grade)
    members_r = await db.execute(
        select(Member).where(Member.status == MemberStatus.ACTIVE, Member.id.notin_(_admin_ids3))
    )
    _all_save = members_r.scalars().all()
    _save_members = (
        _all_save if _mtg_grade == "ALL"
        else [mm for mm in _all_save if _grade_order_s.get(
            mm.masonic_grade.value if hasattr(mm.masonic_grade, "value") else str(mm.masonic_grade), 0
        ) >= _grade_order_s.get(_mtg_grade, 0)]
    )
    for m in _save_members:
        status_val = form.get(f"status_{m.id}", "")
        if not status_val:
            continue
        try:
            new_status = AttendanceStatus(status_val)
        except ValueError:
            continue

        existing_r = await db.execute(
            select(Attendance).where(
                Attendance.meeting_id == meeting_id,
                Attendance.member_id == m.id,
            )
        )
        att = existing_r.scalar_one_or_none()
        excuse = form.get(f"excuse_{m.id}", "").strip() or None

        if att:
            att.status = new_status
            if new_status == AttendanceStatus.EXCUSED:
                att.excuse_reason = excuse
        else:
            db.add(Attendance(
                meeting_id=meeting_id,
                member_id=m.id,
                status=new_status,
                agape=False,
                excuse_reason=excuse if new_status == AttendanceStatus.EXCUSED else None,
            ))

    # ── Maçons passants (présents / no-show) ────────────────────────────────
    mv_r = await db.execute(
        select(MeetingVisitor).where(MeetingVisitor.meeting_id == meeting_id)
    )
    for mv in mv_r.scalars().all():
        mv.status = (
            VisitorStatus.CONFIRMED
            if form.get(f"visitor_present_{mv.id}") == "1"
            else VisitorStatus.CANCELLED
        )

    await db.commit()
    return RedirectResponse(url=f"/attendance/meeting/{meeting_id}?saved=1", status_code=303)


# ── Ajout rapide d'un maçon passant à une tenue ──────────────────────────────

@router.get("/api/visitors/search", response_class=JSONResponse)
async def search_visitors(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    q: str = "",
):
    """Recherche de visiteurs existants pour l'autocomplete (passants assidus)."""
    user, member = ctx
    _require_attendance_mgr(user, member)

    if len(q) < 2:
        # Retourner les visiteurs les plus fréquents
        freq_r = await db.execute(
            select(Visitor, sql_func.count(MeetingVisitor.id).label("visits"))
            .join(MeetingVisitor, MeetingVisitor.visitor_id == Visitor.id)
            .group_by(Visitor.id)
            .order_by(sql_func.count(MeetingVisitor.id).desc())
            .limit(10)
        )
        results = freq_r.all()
    else:
        q_like = f"%{q}%"
        freq_r = await db.execute(
            select(Visitor, sql_func.count(MeetingVisitor.id).label("visits"))
            .join(MeetingVisitor, MeetingVisitor.visitor_id == Visitor.id, isouter=True)
            .where(
                (Visitor.last_name.ilike(q_like)) |
                (Visitor.first_name.ilike(q_like)) |
                (Visitor.lodge_name.ilike(q_like))
            )
            .group_by(Visitor.id)
            .order_by(sql_func.count(MeetingVisitor.id).desc())
            .limit(10)
        )
        results = freq_r.all()

    return JSONResponse([{
        "id": v.id,
        "civility": v.civility or "F",
        "first_name": v.first_name,
        "last_name": v.last_name,
        "lodge_name": v.lodge_name or "",
        "orient_city": v.orient_city or "",
        "obedience": v.obedience or "",
        "is_vm": v.is_vm,
        "visits": visits,
    } for v, visits in results])


@router.post("/meeting/{meeting_id}/add-visitor")
async def add_visitor_to_meeting(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    visitor_id: int = Form(0),
    civility: str = Form("F"),
    first_name: str = Form(""),
    last_name: str = Form(""),
    lodge_name: str = Form(""),
    orient_city: str = Form(""),
    obedience: str = Form(""),
    is_vm: bool = Form(False),
):
    """Ajoute un maçon passant à la tenue (existant ou nouveau)."""
    user, member = ctx
    _require_attendance_mgr(user, member)

    # Vérifier que la tenue existe
    mtg_r = await db.execute(select(Meeting).where(Meeting.id == meeting_id))
    if not mtg_r.scalar_one_or_none():
        raise HTTPException(status_code=404)

    if visitor_id:
        # Visiteur existant
        vis_r = await db.execute(select(Visitor).where(Visitor.id == visitor_id))
        visitor = vis_r.scalar_one_or_none()
        if not visitor:
            raise HTTPException(status_code=404, detail="Visiteur introuvable")
    else:
        # Nouveau visiteur
        if not first_name.strip() or not last_name.strip():
            raise HTTPException(status_code=422, detail="Nom et prénom requis")
        visitor = Visitor(
            civility=civility,
            first_name=first_name.strip(),
            last_name=last_name.strip(),
            lodge_name=lodge_name.strip() or None,
            orient_city=orient_city.strip() or None,
            obedience=obedience.strip() or None,
            is_vm=is_vm,
        )
        db.add(visitor)
        await db.flush()  # pour avoir visitor.id

    # Vérifier qu'il n'est pas déjà inscrit
    existing_r = await db.execute(
        select(MeetingVisitor).where(
            MeetingVisitor.meeting_id == meeting_id,
            MeetingVisitor.visitor_id == visitor.id,
        )
    )
    if not existing_r.scalar_one_or_none():
        db.add(MeetingVisitor(
            meeting_id=meeting_id,
            visitor_id=visitor.id,
            status=VisitorStatus.CONFIRMED,
        ))

    await db.commit()
    return RedirectResponse(url=f"/attendance/meeting/{meeting_id}?saved=1", status_code=303)


# ── Retirer un passant d'une tenue ────────────────────────────────────────────

@router.post("/meeting/{meeting_id}/remove-visitor/{mv_id}")
async def remove_visitor_from_meeting(
    meeting_id: int,
    mv_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    _require_attendance_mgr(user, member)
    mv_r = await db.execute(
        select(MeetingVisitor).where(
            MeetingVisitor.id == mv_id,
            MeetingVisitor.meeting_id == meeting_id,
        )
    )
    mv = mv_r.scalar_one_or_none()
    if mv:
        await db.delete(mv)
        await db.commit()
    return RedirectResponse(url=f"/attendance/meeting/{meeting_id}?saved=1", status_code=303)


# ── Supprimer un visiteur du carnet (toutes tenues) ───────────────────────────

@router.post("/visitor/{visitor_id}/delete")
async def delete_visitor(
    visitor_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    _require_attendance_mgr(user, member)
    vis_r = await db.execute(select(Visitor).where(Visitor.id == visitor_id))
    visitor = vis_r.scalar_one_or_none()
    if visitor:
        # Supprimer d'abord les MeetingVisitor liés
        mv_r = await db.execute(select(MeetingVisitor).where(MeetingVisitor.visitor_id == visitor_id))
        for mv in mv_r.scalars().all():
            await db.delete(mv)
        await db.delete(visitor)
        await db.commit()
    return RedirectResponse(url="/attendance/visitors", status_code=303)


# ── Synthèse post-tenue ───────────────────────────────────────────────────────

@router.get("/meeting/{meeting_id}/summary", response_class=HTMLResponse)
async def meeting_summary(
    request: Request,
    meeting_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    _require_attendance_mgr(user, member)

    result = await db.execute(
        select(Meeting)
        .options(
            selectinload(Meeting.attendances).selectinload(Attendance.member),
            selectinload(Meeting.meeting_visitors).selectinload(MeetingVisitor.visitor),
        )
        .where(Meeting.id == meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404)

    # ── Stats membres ────────────────────────────────────────────────────────
    present = [a for a in meeting.attendances if a.status == AttendanceStatus.PRESENT]
    excused = [a for a in meeting.attendances if a.status == AttendanceStatus.EXCUSED]
    absent  = [a for a in meeting.attendances if a.status == AttendanceStatus.ABSENT]
    total_members = len(meeting.attendances)
    pct_present = round(len(present) * 100 / total_members) if total_members else 0

    # ── Maçons passants confirmés ────────────────────────────────────────────
    confirmed_visitors = [
        mv for mv in meeting.meeting_visitors
        if mv.status == VisitorStatus.CONFIRMED
    ]
    confirmed_visitors.sort(key=lambda mv: mv.visitor.last_name)

    # Loges représentées
    lodges: dict[str, list] = {}
    for mv in confirmed_visitors:
        lodge_key = mv.visitor.lodge_name or "Loge inconnue"
        lodges.setdefault(lodge_key, []).append(mv)

    # Fréquence de visite de chaque passant (toutes tenues confondues)
    visitor_ids = [mv.visitor_id for mv in confirmed_visitors]
    visit_counts = {}   # visitor_id → total visites à notre loge
    if visitor_ids:
        vc_r = await db.execute(
            select(MeetingVisitor.visitor_id, sql_func.count().label("n"))
            .where(MeetingVisitor.visitor_id.in_(visitor_ids))
            .group_by(MeetingVisitor.visitor_id)
        )
        visit_counts = {row.visitor_id: row.n for row in vc_r}

    return templates.TemplateResponse(request, "pages/attendance/summary.html", {
        "current_member": member,
        "current_user": user,
        "meeting": meeting,
        "present": present,
        "excused": excused,
        "absent": absent,
        "total_members": total_members,
        "pct_present": pct_present,
        "confirmed_visitors": confirmed_visitors,
        "lodges": lodges,
        "visit_counts": visit_counts,
    })


# ── Assiduité d'un membre ─────────────────────────────────────────────────────

@router.get("/member/{member_id}", response_class=HTMLResponse)
async def member_attendance(
    request: Request,
    member_id: int,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, current_member = ctx
    if member_id != current_member.id and not (can_manage_attendance(current_member) or user.is_admin):
        raise HTTPException(status_code=403)

    target_r = await db.execute(select(Member).where(Member.id == member_id))
    target = target_r.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404)

    att_r = await db.execute(
        select(Attendance)
        .options(selectinload(Attendance.meeting))
        .where(Attendance.member_id == member_id)
        .order_by(Attendance.meeting_id.desc())
    )
    attendances = att_r.scalars().all()

    total   = len(attendances)
    present = sum(1 for a in attendances if a.status == AttendanceStatus.PRESENT)
    excused = sum(1 for a in attendances if a.status == AttendanceStatus.EXCUSED)
    absent  = sum(1 for a in attendances if a.status == AttendanceStatus.ABSENT)

    return templates.TemplateResponse(request, "pages/attendance/member.html", {
        "current_member": current_member,
        "current_user": user,
        "target": target,
        "attendances": attendances,
        "total": total,
        "present": present,
        "excused": excused,
        "absent": absent,
        "pct": round(present * 100 / total) if total else 0,
    })


# ── Export Excel grille assiduité ────────────────────────────────────────────

@router.get("/export/csv")
async def attendance_export_excel(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    year_id: int = 0,
):
    user, member = ctx
    _require_attendance_mgr(user, member)

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    years_r = await db.execute(
        select(MasonicYear).order_by(MasonicYear.is_current.desc(), MasonicYear.start_date.desc())
    )
    years = years_r.scalars().all()
    current_year = next((y for y in years if y.is_current), years[0] if years else None)
    selected_year = next((y for y in years if y.id == year_id), current_year)

    year_filter = (Meeting.masonic_year_id == selected_year.id) if selected_year else True

    past_r = await db.execute(
        select(Meeting)
        .where(year_filter, Meeting.meeting_date <= date.today())
        .order_by(Meeting.meeting_date)
    )
    past_meetings = past_r.scalars().all()
    past_ids = [m.id for m in past_meetings]

    from app.models.identity import User as _UserExp
    _admin_exp_r = await db.execute(
        select(_UserExp.member_id).where(_UserExp.is_admin == True, _UserExp.member_id.isnot(None))
    )
    _admin_exp = {row[0] for row in _admin_exp_r}
    members_r = await db.execute(
        select(Member)
        .where(Member.status == MemberStatus.ACTIVE, Member.id.notin_(_admin_exp))
        .order_by(Member.last_name, Member.first_name)
    )
    active_members = members_r.scalars().all()

    # Seuils
    ls_r = await db.execute(select(LodgeSettings).limit(1))
    lodge = ls_r.scalar_one_or_none()
    threshold_warn   = (lodge.attendance_threshold_warn   if lodge else None) or 70
    threshold_danger = (lodge.attendance_threshold_danger if lodge else None) or 50

    grid: dict[int, dict[int, str]] = {}
    if past_ids:
        att_r = await db.execute(
            select(Attendance).where(Attendance.meeting_id.in_(past_ids))
        )
        for att in att_r.scalars().all():
            grid.setdefault(att.member_id, {})[att.meeting_id] = att.status.value

    # Couleurs
    thin = Side(style="thin", color="D1D5DB")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    F_PRESENT = PatternFill("solid", fgColor="DCFCE7")
    F_EXCUSED = PatternFill("solid", fgColor="FEF9C3")
    F_ABSENT  = PatternFill("solid", fgColor="FEE2E2")
    F_EMPTY   = PatternFill("solid", fgColor="F9FAFB")
    F_HEADER  = PatternFill("solid", fgColor="374151")
    F_NAME    = PatternFill("solid", fgColor="F1F5F9")
    F_GREEN   = PatternFill("solid", fgColor="DCFCE7")
    F_WARN    = PatternFill("solid", fgColor="FEF9C3")
    F_DANGER  = PatternFill("solid", fgColor="FEE2E2")
    F_TOTAL   = PatternFill("solid", fgColor="E0E7FF")

    wb = Workbook()

    # ── Feuille 1 : Grille de présence ───────────────────────────────────────
    ws = wb.active
    ws.title = "Grille présences"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C5"  # fige nom/prénom + ligne d'en-tête

    # Titre
    last_col = get_column_letter(2 + len(past_meetings) + 4)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Grille d'assiduité — {selected_year.label if selected_year else ''}"
    ws["A1"].font  = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill  = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{len(past_meetings)} tenue(s) · {len(active_members)} membre(s) actif(s) · Exporté le {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font  = Font(italic=True, size=9, color="9CA3AF")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Légende (ligne 3)
    for col, (label, fill) in enumerate([
        ("P = Présent", F_PRESENT), ("E = Excusé", F_EXCUSED),
        ("A = Absent", F_ABSENT), ("· = Non enregistré", F_EMPTY)
    ], 1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = fill
        cell.font = Font(size=8, color="374151")
        cell.alignment = Alignment(horizontal="center")

    # En-têtes colonnes (ligne 4)
    ws.cell(row=4, column=1, value="Nom").fill   = F_HEADER
    ws.cell(row=4, column=1).font  = Font(bold=True, color="FFFFFF")
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=4, column=2, value="Prénom").fill = F_HEADER
    ws.cell(row=4, column=2).font  = Font(bold=True, color="FFFFFF")
    ws.cell(row=4, column=2).alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18

    for i, m in enumerate(past_meetings):
        col = 3 + i
        cell = ws.cell(row=4, column=col,
                       value=m.meeting_date.strftime("%d/%m"))
        cell.fill = F_HEADER
        cell.font = Font(bold=True, color="FFFFFF", size=8)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 6

    stat_col = 3 + len(past_meetings)
    for c, label in enumerate(["P", "E", "A", "%"], stat_col):
        cell = ws.cell(row=4, column=c, value=label)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = 6

    ws.row_dimensions[4].height = 20

    # Tenues applicables par membre (grade + date d'arrivée)
    _grade_order_exp = {"APPRENTI": 1, "COMPAGNON": 2, "MAITRE": 3, "ALL": 0}
    F_NA = PatternFill("solid", fgColor="E5E7EB")  # gris = non concerné

    # Données membres
    for row_idx, mbr in enumerate(active_members):
        r = 5 + row_idx
        member_grid = grid.get(mbr.id, {})
        present = excused = absent = 0
        grade_val = mbr.masonic_grade.value if hasattr(mbr.masonic_grade, "value") else str(mbr.masonic_grade)
        start = mbr.membership_start_date

        ws.cell(row=r, column=1, value=mbr.last_name).fill  = F_NAME
        ws.cell(row=r, column=1).border = brd
        ws.cell(row=r, column=1).font = Font(bold=True, size=9)
        ws.cell(row=r, column=2, value=f"{'S∴' if mbr.civility == 'S' else 'F∴'} {mbr.first_name}").fill = F_NAME
        ws.cell(row=r, column=2).border = brd
        ws.cell(row=r, column=2).font = Font(size=9)

        expected = 0
        for i, m in enumerate(past_meetings):
            col = 3 + i
            mtg_grade = m.grade.value if hasattr(m.grade, "value") else str(m.grade)
            applicable = (
                (not start or m.meeting_date >= start)
                and (mtg_grade == "ALL" or _grade_order_exp.get(grade_val, 0) >= _grade_order_exp.get(mtg_grade, 0))
            )
            val = member_grid.get(m.id, "")
            if not applicable:
                label, fill, bold, color = "—", F_NA, False, "9CA3AF"
            elif val == "PRESENT":
                label, fill, bold, color = "P", F_PRESENT, True, "166534"
                present += 1
                expected += 1
            elif val == "EXCUSED":
                label, fill, bold, color = "E", F_EXCUSED, False, "92400E"
                excused += 1
                expected += 1
            elif val == "ABSENT":
                label, fill, bold, color = "A", F_ABSENT, True, "991B1B"
                absent += 1
                expected += 1
            else:
                label, fill, bold, color = "·", F_EMPTY, False, "D1D5DB"
                expected += 1
            cell = ws.cell(row=r, column=col, value=label)
            cell.fill = fill
            cell.border = brd
            cell.font = Font(bold=bold, size=9, color=color)
            cell.alignment = Alignment(horizontal="center")

        pct = round(present * 100 / expected) if expected else 0
        pct_fill = F_GREEN if pct >= threshold_warn else F_WARN if pct >= threshold_danger else F_DANGER

        for c, val in enumerate([present, excused, absent, f"{pct}%"], stat_col):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = brd
            cell.font = Font(bold=(c == stat_col + 3), size=9)
            cell.alignment = Alignment(horizontal="center")
            if c == stat_col + 3:
                cell.fill = pct_fill
            else:
                fills_stat = [F_PRESENT, F_EXCUSED, F_ABSENT]
                cell.fill = fills_stat[c - stat_col]

        ws.row_dimensions[r].height = 16

    # ── Feuille 2 : Statistiques individuelles ────────────────────────────────
    ws2 = wb.create_sheet("Stats membres")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"Statistiques d'assiduité — {selected_year.label if selected_year else ''}"
    ws2["A1"].font  = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill  = PatternFill("solid", fgColor="1E3A5F")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for col, label in enumerate(["Nom", "Prénom", "Présents", "Excusés", "Absents",
                                  "Total enreg.", "Assiduité %"], 1):
        cell = ws2.cell(row=3, column=col, value=label)
        cell.fill  = F_HEADER
        cell.font  = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = brd

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20
    for col in "CDEFG":
        ws2.column_dimensions[col].width = 14

    for row_idx, mbr in enumerate(active_members):
        r = 4 + row_idx
        member_grid = grid.get(mbr.id, {})
        grade_val2 = mbr.masonic_grade.value if hasattr(mbr.masonic_grade, "value") else str(mbr.masonic_grade)
        start2 = mbr.membership_start_date
        present = excused = absent = expected2 = 0
        for m in past_meetings:
            mtg_grade2 = m.grade.value if hasattr(m.grade, "value") else str(m.grade)
            if (not start2 or m.meeting_date >= start2) and (
                mtg_grade2 == "ALL" or _grade_order_exp.get(grade_val2, 0) >= _grade_order_exp.get(mtg_grade2, 0)
            ):
                expected2 += 1
                v = member_grid.get(m.id, "")
                if v == "PRESENT":   present += 1
                elif v == "EXCUSED": excused += 1
                elif v == "ABSENT":  absent  += 1
        total   = present + excused + absent
        pct     = round(present * 100 / expected2) if expected2 else 0
        pct_fill = F_GREEN if pct >= threshold_warn else F_WARN if pct >= threshold_danger else F_DANGER

        row_fill = F_NAME
        for c, val in enumerate([mbr.last_name,
                                  f"{'S∴' if mbr.civility == 'S' else 'F∴'} {mbr.first_name}",
                                  present, excused, absent, total, f"{pct}%"], 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = brd
            cell.font   = Font(size=10, bold=(c == 7))
            cell.alignment = Alignment(horizontal="center" if c > 2 else "left",
                                       indent=1 if c <= 2 else 0)
            cell.fill = pct_fill if c == 7 else row_fill
        ws2.row_dimensions[r].height = 18

    # ── Feuille 3 : Vue synthèse par tenue ────────────────────────────────────
    ws3 = wb.create_sheet("Par tenue")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    ws3["A1"] = "Synthèse par tenue"
    ws3["A1"].font  = Font(bold=True, size=13, color="FFFFFF")
    ws3["A1"].fill  = PatternFill("solid", fgColor="1E3A5F")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    for col, label in enumerate(["Date", "Type", "Présents", "Excusés", "Absents"], 1):
        cell = ws3.cell(row=3, column=col, value=label)
        cell.fill = F_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = brd

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 12

    # Calculer stats par tenue
    meeting_stats: dict[int, dict] = {}
    for att_row in grid.values():
        for mid, status in att_row.items():
            s = meeting_stats.setdefault(mid, {"P": 0, "E": 0, "A": 0})
            if status == "PRESENT":   s["P"] += 1
            elif status == "EXCUSED": s["E"] += 1
            elif status == "ABSENT":  s["A"] += 1

    from app.models.meetings import MeetingType as _MT
    type_labels = {
        "BLANCHE": "Tenue blanche", "SOLENNELLE": "Tenue solennelle",
        "INSTRUCTION": "Tenue d'instruction", "INITIATION": "Initiation",
        "INSTALLATION": "Installation", "ELECTION": "Élection",
        "PASSAGE": "Passage 2e degré", "ELEVATION": "Élévation 3e degré",
        "FETE": "Fête maçonnique", "EXTRA": "Tenue extraordinaire",
    }
    for row_idx, m in enumerate(past_meetings):
        r = 4 + row_idx
        s = meeting_stats.get(m.id, {"P": 0, "E": 0, "A": 0})
        total = s["P"] + s["E"] + s["A"]
        for col, val in enumerate([
            m.meeting_date.strftime("%d/%m/%Y"),
            m.title or type_labels.get(m.type.value, m.type.value),
            s["P"], s["E"], s["A"]
        ], 1):
            cell = ws3.cell(row=r, column=col, value=val)
            cell.border = brd
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left",
                                       indent=1 if col <= 2 else 0)
            fill_map = {3: F_PRESENT, 4: F_EXCUSED, 5: F_ABSENT}
            cell.fill = fill_map.get(col, PatternFill("solid", fgColor="FFFFFF"))
        ws3.row_dimensions[r].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    year_label = selected_year.label.replace(" ", "_") if selected_year else "export"
    filename = f"assiduite_{year_label}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Dashboard maçons passants ─────────────────────────────────────────────────

@router.get("/visitors", response_class=HTMLResponse)
async def visitors_dashboard(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user, member = ctx
    _require_attendance_mgr(user, member)

    vis_r = await db.execute(
        select(
            Visitor,
            sql_func.count(MeetingVisitor.id).label("visit_count"),
            sql_func.max(Meeting.meeting_date).label("last_visit"),
        )
        .join(MeetingVisitor, MeetingVisitor.visitor_id == Visitor.id, isouter=True)
        .join(Meeting, Meeting.id == MeetingVisitor.meeting_id, isouter=True)
        .where(
            (MeetingVisitor.status == VisitorStatus.CONFIRMED) |
            (MeetingVisitor.id.is_(None))
        )
        .group_by(Visitor.id)
        .order_by(sql_func.count(MeetingVisitor.id).desc(), Visitor.last_name)
    )
    visitor_rows = vis_r.all()

    total_visits  = sum(r.visit_count or 0 for r in visitor_rows)
    total_unique  = len(visitor_rows)
    optin_count   = sum(1 for r in visitor_rows if r.Visitor.program_optin)

    lodge_r = await db.execute(
        select(Visitor.lodge_name, sql_func.count().label("n"))
        .join(MeetingVisitor, MeetingVisitor.visitor_id == Visitor.id)
        .where(MeetingVisitor.status == VisitorStatus.CONFIRMED, Visitor.lodge_name.isnot(None))
        .group_by(Visitor.lodge_name).order_by(sql_func.count().desc()).limit(15)
    )
    top_lodges = lodge_r.all()

    orient_r = await db.execute(
        select(Visitor.orient_city, sql_func.count().label("n"))
        .join(MeetingVisitor, MeetingVisitor.visitor_id == Visitor.id)
        .where(MeetingVisitor.status == VisitorStatus.CONFIRMED, Visitor.orient_city.isnot(None))
        .group_by(Visitor.orient_city).order_by(sql_func.count().desc()).limit(15)
    )
    top_orients = orient_r.all()

    obed_r = await db.execute(
        select(Visitor.obedience, sql_func.count().label("n"))
        .join(MeetingVisitor, MeetingVisitor.visitor_id == Visitor.id)
        .where(MeetingVisitor.status == VisitorStatus.CONFIRMED, Visitor.obedience.isnot(None))
        .group_by(Visitor.obedience).order_by(sql_func.count().desc()).limit(10)
    )
    top_obediences = obed_r.all()

    return templates.TemplateResponse(request, "pages/attendance/visitors.html", {
        "current_member": member,
        "current_user": user,
        "visitor_rows": visitor_rows,
        "total_visits": total_visits,
        "total_unique": total_unique,
        "optin_count": optin_count,
        "top_lodges": top_lodges,
        "top_orients": top_orients,
        "top_obediences": top_obediences,
    })


# ── Bilan d'activité annuel ───────────────────────────────────────────────────

@router.get("/bilan", response_class=HTMLResponse)
async def bilan_annuel(
    request: Request,
    ctx: Annotated[tuple, Depends(require_auth)],
    db: Annotated[AsyncSession, Depends(get_db)],
    year_id: int = 0,
    date_from: str = "",
    date_to: str = "",
):
    user, member = ctx
    _require_attendance_mgr(user, member)

    years_r = await db.execute(
        select(MasonicYear).order_by(MasonicYear.is_current.desc(), MasonicYear.start_date.desc())
    )
    years = years_r.scalars().all()
    current_year = next((y for y in years if y.is_current), years[0] if years else None)
    selected_year = next((y for y in years if y.id == year_id), current_year)

    from datetime import date as _date, datetime as _dt
    if date_from and date_to:
        try:
            d_from = _dt.strptime(date_from, "%Y-%m-%d").date()
            d_to   = _dt.strptime(date_to,   "%Y-%m-%d").date()
        except ValueError:
            d_from = selected_year.start_date if selected_year else _date.today().replace(month=9, day=1)
            d_to   = selected_year.end_date   if selected_year else _date.today()
    elif selected_year:
        d_from = selected_year.start_date
        d_to   = selected_year.end_date or _date.today()
    else:
        d_from = _date.today().replace(month=9, day=1)
        d_to   = _date.today()

    mtg_r = await db.execute(
        select(Meeting)
        .where(Meeting.meeting_date.between(d_from, d_to))
        .order_by(Meeting.meeting_date)
    )
    meetings = mtg_r.scalars().all()
    past_meetings = [m for m in meetings if m.meeting_date <= _date.today()]
    past_ids = [m.id for m in past_meetings]

    from app.models.identity import User as _UserBilan
    _adm_r = await db.execute(
        select(_UserBilan.member_id).where(_UserBilan.is_admin == True, _UserBilan.member_id.isnot(None))
    )
    _adm_ids = {row[0] for row in _adm_r}
    mem_r = await db.execute(
        select(Member)
        .where(Member.status == MemberStatus.ACTIVE, Member.id.notin_(_adm_ids))
        .order_by(Member.last_name, Member.first_name)
    )
    active_members = mem_r.scalars().all()

    grid2: dict[int, dict[int, str]] = {}
    if past_ids:
        att_r = await db.execute(select(Attendance).where(Attendance.meeting_id.in_(past_ids)))
        for att in att_r.scalars().all():
            grid2.setdefault(att.member_id, {})[att.meeting_id] = att.status.value

    _go = {"APPRENTI": 1, "COMPAGNON": 2, "MAITRE": 3, "ALL": 0}
    member_stats = []
    g_exp = g_pres = g_exc = g_abs = 0
    for mbr in active_members:
        gv    = mbr.masonic_grade.value if hasattr(mbr.masonic_grade, "value") else str(mbr.masonic_grade)
        start = mbr.membership_start_date
        mg    = grid2.get(mbr.id, {})
        exp = pres = exc = ab = 0
        for m in past_meetings:
            if start and m.meeting_date < start:
                continue
            mgv = m.grade.value if hasattr(m.grade, "value") else str(m.grade)
            if mgv != "ALL" and _go.get(gv, 0) < _go.get(mgv, 0):
                continue
            exp += 1
            v = mg.get(m.id, "")
            if v == "PRESENT":   pres += 1
            elif v == "EXCUSED": exc  += 1
            elif v == "ABSENT":  ab   += 1
        g_exp += exp; g_pres += pres; g_exc += exc; g_abs += ab
        member_stats.append({
            "member": mbr, "expected": exp, "present": pres,
            "excused": exc, "absent": ab,
            "pct": round(pres * 100 / exp) if exp else 0,
        })

    from collections import Counter
    type_counter = Counter(m.type.value for m in past_meetings)

    mv_r = await db.execute(
        select(MeetingVisitor)
        .options(selectinload(MeetingVisitor.visitor))
        .where(
            MeetingVisitor.meeting_id.in_(past_ids) if past_ids else False,
            MeetingVisitor.status == VisitorStatus.CONFIRMED,
        )
    )
    all_mvs = mv_r.scalars().all() if past_ids else []
    visitor_total  = len(all_mvs)
    visitor_unique = len({mv.visitor_id for mv in all_mvs})
    top_lodges_bilan  = Counter(mv.visitor.lodge_name  for mv in all_mvs if mv.visitor.lodge_name).most_common(10)
    top_orients_bilan = Counter(mv.visitor.orient_city for mv in all_mvs if mv.visitor.orient_city).most_common(5)

    # Top 5 visiteurs les plus assidus sur la période
    visitor_visit_counts: Counter = Counter()
    visitor_by_id: dict = {}
    for mv in all_mvs:
        visitor_visit_counts[mv.visitor_id] += 1
        visitor_by_id[mv.visitor_id] = mv.visitor
    top_visitors = [
        {"visitor": visitor_by_id[vid], "visits": cnt}
        for vid, cnt in visitor_visit_counts.most_common(5)
    ]

    # Stats présences par tenue (depuis grid2)
    mtg_att_stats: dict[int, dict] = {}
    for mbr_id, mbr_grid in grid2.items():
        for mid, status in mbr_grid.items():
            s = mtg_att_stats.setdefault(mid, {"P": 0, "E": 0, "A": 0})
            if status == "PRESENT":   s["P"] += 1
            elif status == "EXCUSED": s["E"] += 1
            elif status == "ABSENT":  s["A"] += 1

    # Nombre de passants par tenue
    mtg_visitor_counts = Counter(mv.meeting_id for mv in all_mvs)

    # ── Travaux / planches par tenue ────────────────────────────────────────
    mtg_travaux: dict[int, str] = {}
    mtg_odj: dict[int, str] = {}
    if past_ids:
        ts_r = await db.execute(
            select(TracingSection).where(
                TracingSection.meeting_id.in_(past_ids),
                TracingSection.section_type.in_([TracingSectionType.TRAVAUX, TracingSectionType.ODJ]),
            )
        )
        for ts in ts_r.scalars().all():
            if ts.section_type == TracingSectionType.TRAVAUX and ts.content_html:
                mtg_travaux[ts.meeting_id] = ts.content_html
            elif ts.section_type == TracingSectionType.ODJ and ts.content_html:
                mtg_odj[ts.meeting_id] = ts.content_html

    # ── Agapes par tenue ────────────────────────────────────────────────────
    # Membres avec agape=True + leurs invités
    mtg_agape: dict[int, dict] = {}
    if past_ids:
        agape_r = await db.execute(
            select(Attendance).where(
                Attendance.meeting_id.in_(past_ids),
                Attendance.agape == True,
            )
        )
        for att in agape_r.scalars().all():
            s = mtg_agape.setdefault(att.meeting_id, {"members": 0, "guests": 0})
            s["members"] += 1
            s["guests"] += att.agape_guests or 0
    # Passants avec agape=True
    for mv in all_mvs:
        if mv.agape:
            s = mtg_agape.setdefault(mv.meeting_id, {"members": 0, "guests": 0})
            s["guests"] += 1 + (mv.agape_guests or 0)

    total_agape_covers = sum(s["members"] + s["guests"] for s in mtg_agape.values())
    mtg_with_agape = len(mtg_agape)

    type_labels_b = {
        "BLANCHE": "Tenue blanche", "SOLENNELLE": "Tenue solennelle",
        "INSTRUCTION": "Tenue d'instruction", "INITIATION": "Initiation",
        "INSTALLATION": "Installation", "ELECTION": "Election",
        "PASSAGE": "Passage 2e degre", "ELEVATION": "Elevation 3e degre",
        "FETE": "Fete maconnique", "EXTRA": "Tenue extraordinaire",
    }

    ls_r = await db.execute(select(LodgeSettings).limit(1))
    lodge_settings = ls_r.scalar_one_or_none()
    lodge_name_b = lodge_settings.name if lodge_settings else "Loge"

    return templates.TemplateResponse(request, "pages/attendance/bilan.html", {
        "current_member": member,
        "current_user": user,
        "years": years,
        "selected_year": selected_year,
        "d_from": d_from,
        "d_to": d_to,
        "meetings": meetings,
        "past_meetings": past_meetings,
        "active_members": active_members,
        "member_stats": member_stats,
        "g_exp": g_exp, "g_pres": g_pres, "g_exc": g_exc, "g_abs": g_abs,
        "g_pct_pres": round(g_pres * 100 / g_exp) if g_exp else 0,
        "g_pct_exc":  round(g_exc  * 100 / g_exp) if g_exp else 0,
        "g_pct_abs":  round(g_abs  * 100 / g_exp) if g_exp else 0,
        "type_counter": dict(type_counter),
        "type_labels": type_labels_b,
        "visitor_total": visitor_total,
        "visitor_unique": visitor_unique,
        "top_lodges_bilan": top_lodges_bilan,
        "top_orients_bilan": top_orients_bilan,
        "top_visitors": top_visitors,
        "lodge_name": lodge_name_b,
        "mtg_att_stats": mtg_att_stats,
        "mtg_visitor_counts": dict(mtg_visitor_counts),
        "mtg_travaux": mtg_travaux,
        "mtg_odj": mtg_odj,
        "mtg_agape": mtg_agape,
        "total_agape_covers": total_agape_covers,
        "mtg_with_agape": mtg_with_agape,
    })
