"""
Import tenues + visiteurs depuis le bilan Excel de l'ancien outil.
Usage : python import_tenues.py <chemin_excel>
"""
import re
import sys
import sqlite3
from datetime import datetime

DB_PATH  = "socrate_local.db"
YEAR_ID  = 1   # année maçonnique 2025-2026

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl requis : pip install openpyxl")

if len(sys.argv) < 2:
    sys.exit(f"Usage: python {sys.argv[0]} <fichier.xlsx>")

EXCEL = sys.argv[1]

# ── helpers ──────────────────────────────────────────────────────────────────

def to_date(v) -> str:
    if isinstance(v, (datetime,)):
        return v.strftime("%Y-%m-%d")
    if hasattr(v, "strftime"):          # date object
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]

def detect_type(title: str) -> str:
    t = (title or "").lower()
    if any(k in t for k in ("réception", "reception", "initiation")):
        return "INITIATION"
    if any(k in t for k in ("passage", "chambre de compagnon")):
        return "PASSAGE"
    if any(k in t for k in ("élévation", "elevation", "maîtrise", "maitrise")):
        return "ELEVATION"
    if "installation" in t:
        return "INSTALLATION"
    if any(k in t for k in ("sol∴", "solennelle", "sol:")):
        return "SOLENNELLE"
    return "BLANCHE"

def extract_number(title: str):
    m = re.match(r"(\d+)", title or "")
    return int(m.group(1)) if m else None

# ── connexion ─────────────────────────────────────────────────────────────────

con = sqlite3.connect(DB_PATH)
cur = con.cursor()
cur.execute("PRAGMA foreign_keys = ON")

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ── 1. TENUES ─────────────────────────────────────────────────────────────────

print("\n=== Import des tenues ===")

# Index des dates déjà présentes pour cette année
cur.execute("SELECT meeting_date, id FROM meetings WHERE masonic_year_id = ?", (YEAR_ID,))
date_to_id: dict[str, int] = {r[0]: r[1] for r in cur.fetchall()}

GRADE_MAP = {"APPRENTI": "APPRENTI", "COMPAGNON": "COMPAGNON", "MAITRE": "MAITRE"}

imported_meetings = 0
skipped_meetings  = 0

ws = wb["Tenues"]
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    date_val, grade_str, title, locked_str, _visitors = row

    date_str  = to_date(date_val)
    grade     = GRADE_MAP.get(grade_str, "APPRENTI")
    is_locked = 1 if str(locked_str or "").strip().lower() == "oui" else 0
    mtype     = detect_type(title)
    mnum      = extract_number(title)

    if date_str in date_to_id:
        print(f"  ↷ Ignorée (déjà en base) : {date_str}  {title[:60]}")
        skipped_meetings += 1
        continue

    cur.execute("""
        INSERT INTO meetings
            (masonic_year_id, meeting_date, grade, type, title, meeting_number,
             is_locked, registration_open, agape_enabled,
             token)
        VALUES (?, ?, ?, ?, ?, ?, ?, 0, 1,
                lower(hex(randomblob(32))))
    """, (YEAR_ID, date_str, grade, mtype, title, mnum, is_locked))

    new_id = cur.lastrowid
    date_to_id[date_str] = new_id
    imported_meetings += 1
    print(f"  ✓ Créée : {date_str}  [{grade}] #{mnum}  {title[:60]}")

con.commit()
print(f"\n  → {imported_meetings} tenues importées, {skipped_meetings} ignorées.")

# ── 2. VISITEURS ──────────────────────────────────────────────────────────────

print("\n=== Import des visiteurs ===")

imported_visitors = 0
linked_visitors   = 0
skipped_mv        = 0

ws2 = wb["Visiteurs"]
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    date_val, nom, prenom, loge, orient, obedience, email, optin_str = row

    date_str   = to_date(date_val)
    meeting_id = date_to_id.get(date_str)
    if not meeting_id:
        print(f"  ⚠ Tenue introuvable pour la date {date_str} ({nom} {prenom})")
        continue

    nom    = (nom    or "").strip()
    prenom = (prenom or "").strip()

    # Trouver ou créer le visiteur
    cur.execute("""
        SELECT id FROM visitors
        WHERE upper(last_name) = upper(?) AND upper(first_name) = upper(?)
    """, (nom, prenom))
    row_v = cur.fetchone()

    if row_v:
        visitor_id = row_v[0]
    else:
        optin = 1 if str(optin_str or "").strip().lower() == "oui" else 0
        cur.execute("""
            INSERT INTO visitors
                (last_name, first_name, lodge_name, orient_city, obedience, email, program_optin)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (nom, prenom, loge or "", orient or "", obedience or "",
              email if email else None, optin))
        visitor_id = cur.lastrowid
        imported_visitors += 1

    # Lier à la tenue (évite les doublons)
    cur.execute("""
        SELECT id FROM meeting_visitors WHERE meeting_id = ? AND visitor_id = ?
    """, (meeting_id, visitor_id))
    if cur.fetchone():
        skipped_mv += 1
        continue

    cur.execute("""
        INSERT INTO meeting_visitors (meeting_id, visitor_id, status, agape, agape_guests)
        VALUES (?, ?, 'CONFIRMED', 0, 0)
    """, (meeting_id, visitor_id))
    linked_visitors += 1
    print(f"  ✓ {date_str}  {nom} {prenom}  ({loge} — {orient})")

con.commit()
con.close()

print(f"\n  → {imported_visitors} nouveaux visiteurs, {linked_visitors} liens tenue-visiteur créés, {skipped_mv} doublons ignorés.")
print("\n✅ Import terminé.")
